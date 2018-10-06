#!/usr/bin/env python

from sys import argv
from pprint import pprint
from openpyxl import Workbook, load_workbook

from libs.transformations import COLUMN_TRANSF_RULES
from libs.color import get_color
from libs.utilities import *

# Global Data
MODEL_INDX  = 0
MANUF_INDX  = 1
CATEG_INDX  = 2
COLLEC_INDX = 3
FAMILY_INDX = 4
GENDER_INDX = 5
PRICE_INDX  = 6
DISC_INDX   = 7
QUANT_INDX  = 8
HIDDN_INDX  = 9
IMG_INDX    = 10

MANUFACTURER = ['Chronoswiss',
                'Fortis',
                'Jos Von Arx',
                'Jowissa',
                'Manfred Cracco',
                'Rodania',
                'Victorinox']

def cleanup(wb):
    products_sheet = wb['Products']
    products_sheet.append(['' for i in range(products_sheet.max_column)])
    row_num = products_sheet.max_row

    # print("Row to clean is: ", row_num)
    products_sheet.delete_rows(row_num)

def process_attr_data(product, attr_grp):
    
    for attr in COLUMN_TRANSF_RULES[attr_grp]:
        # First create the list(of lists) based on the priorities
        transformations_ordered = [[] for i in range(len(COLUMN_TRANSF_RULES[attr_grp][attr]) + 1000)] 
        
        for word, transf in COLUMN_TRANSF_RULES[attr_grp][attr].items():
            transformations_ordered[transf[-1]].append(word)
            # [ ['STEEL','S.S'], ['Ceramic'], ... ]
            
        # Clean empty lists
        transformations_ordered = [lst for lst in transformations_ordered if lst]
        # if (attr == 'ΥΛΙΚΟ ΔΕΣΙΜΑΤΟΣ'):
        #     pprint(transformations_ordered)
        # Watch out for no data
        try:
            if not product[attr]:
                product[attr] = ""
                continue
        except:
            product[attr] = ""


        new_val_el = ""
        new_val_en = ""
        for transf in transformations_ordered:
            possible_words = [word for word in transf]

            max_word = ""
            # Now check if any of the words are in the text
            for word in possible_words:

                if "__color" in word:
                    
                    colors_vals = get_color(word.lower(), replace_chars("/|", product[attr].lower(), " "))
                    if colors_vals:
                        if len(word) > len(max_word):
                            max_word = word

                elif word.upper() in product[attr].upper():
                    if len(word) > len(max_word):
                        max_word = word

            # If a word was matched, it will be the one with the most characters
            if max_word:
                new_val_en += ' ' + COLUMN_TRANSF_RULES[attr_grp][attr][max_word][0]
                new_val_el += ' ' + COLUMN_TRANSF_RULES[attr_grp][attr][max_word][1]

                # If the pattern matched we have the colors that need to be replaced
                if "__color" in max_word:
                    colors_vals = get_color(max_word.lower(), replace_chars("/|", product[attr].lower(), " "))
                    new_val_el = new_val_el.replace("__color", colors_vals[1])
                    new_val_en = new_val_en.replace("__color", colors_vals[0])
        
        if new_val_el == "":
            product[attr] = [product[attr], product[attr]]
        else:
            new_val_el = new_val_el[1:][0].upper() + new_val_el[2:]
            new_val_en = new_val_en[1:][0].upper() + new_val_en[2:]
            if new_val_el[-1] == ',':
                new_val_el = new_val_el[:-1]
                new_val_en = new_val_en[:-1]
            product[attr] = [new_val_el, new_val_en]

    # Lastly all attributes that weren't filtered, convert them to [greek, engl] format
    for attr in product:
        if not isinstance(product[attr], list):
            product[attr] = [product[attr], product[attr]]

    return product        

def static_pre_processing(product_info, attribute_info, attr_grp):

    # Apply rules to the data before inserting it
    attribute_info['ΣΥΛΛΟΓΗ'] = product_info[COLLEC_INDX] + ' ' + product_info[FAMILY_INDX]
    if attribute_info['ΔΙΑΜΕΤΡΟΣ ΚΑΣΑΣ'] != "":
        attribute_info['ΔΙΑΜΕΤΡΟΣ ΚΑΣΑΣ'] = attribute_info['ΔΙΑΜΕΤΡΟΣ ΚΑΣΑΣ'] + "mm"
    if attribute_info['ΠΑΧΟΣ ΚΑΣΑΣ'] != "":
        attribute_info['ΠΑΧΟΣ ΚΑΣΑΣ'] = attribute_info['ΠΑΧΟΣ ΚΑΣΑΣ'] + "mm"

    return attribute_info

def static_post_processing(product_info, attribute_info, attr_grp):

    # ΔΕΣΙΜΟ
    if attribute_info['ΔΕΣΙΜΟ'] == ["",""]:
        return attribute_info

    if attribute_info['ΔΕΣΙΜΟ'][0] == "bracelet":
        en_val = attribute_info['ΥΛΙΚΟ ΔΕΣΙΜΑΤΟΣ'][1] + ' bracelet'
        en_val = rmcomma(en_val, " and", 2)
        gr_val = 'Μπρασελέ από ' + attribute_info['ΥΛΙΚΟ ΔΕΣΙΜΑΤΟΣ'][0].lower()
        gr_val = rmcomma(gr_val, " και")
    elif attribute_info['ΔΕΣΙΜΟ'][0] == "strap":
        en_val = attribute_info['ΥΛΙΚΟ ΔΕΣΙΜΑΤΟΣ'][1] + ' strap'
        en_val = rmcomma(en_val, " and", 2)
        gr_val = attribute_info['ΥΛΙΚΟ ΔΕΣΙΜΑΤΟΣ'][0] + ' λουράκι'
        en_val = rmcomma(en_val, " και", 2)

    attribute_info['ΥΛΙΚΟ ΔΕΣΙΜΑΤΟΣ'][0] = gr_val
    attribute_info['ΥΛΙΚΟ ΔΕΣΙΜΑΤΟΣ'][1] = en_val

    attribute_info['ΔΕΣΙΜΟ'] = ["",""]

    # ΛΕΙΤΟΥΡΓΙΕΣ
    attribute_info['ΛΕΙΤΟΥΡΓΙΕΣ'][0] = rmcomma(attribute_info['ΛΕΙΤΟΥΡΓΙΕΣ'][0], " και", 2)
    attribute_info['ΛΕΙΤΟΥΡΓΙΕΣ'][1] = rmcomma(attribute_info['ΛΕΙΤΟΥΡΓΙΕΣ'][1], " and", 2)

    # ΚΑΣΑ
    attribute_info['ΚΑΣΑ'][0] = rmcomma(attribute_info['ΚΑΣΑ'][0], " και", 2)
    attribute_info['ΚΑΣΑ'][1] = rmcomma(attribute_info['ΚΑΣΑ'][1], " and", 2)

    return attribute_info

def open_new_products(input_file):
    new_products = open(input_file)
    new_products = [line.replace('\n','') for line in new_products]

    # Remove commas and trim spaces
    new_products = [line.split(',') for line in new_products]

    for product in range(len(new_products)):
        for entry in range(len(new_products[product])):
            new_products[product][entry] = \
                                        new_products[product][entry].strip()

    # Format data
    for product in range(len(new_products)):
        # Manufacturer
        new_products[product][MANUF_INDX] = \
                                   new_products[product][MANUF_INDX].title()
        # Category
        new_products[product][CATEG_INDX] = \
                          new_products[product][CATEG_INDX].replace(' ', '')
        new_products[product][CATEG_INDX] = \
                                   new_products[product][CATEG_INDX].upper()
        # Gender
        new_products[product][GENDER_INDX] = \
                                  new_products[product][GENDER_INDX].lower()
        # Price
        new_products[product][PRICE_INDX] = \
                                           new_products[product][PRICE_INDX]

        # Family
        new_products[product][FAMILY_INDX] = \
                                  new_products[product][FAMILY_INDX].upper()

        # Collection
        new_products[product][COLLEC_INDX] = \
                                  new_products[product][COLLEC_INDX].title()

    
    return new_products

def open_product_attributes(input_file):
    new_attrs = open(input_file)
    new_attrs = [line.replace('\n','') for line in new_attrs] 
    new_attrs = [line.split(',') for line in new_attrs]
    attr_names = new_attrs.pop(0)

    attrs_dicts = [{} for prod in new_attrs]
    for prod_ind in range(len(new_attrs)):
        for i in range(len(new_attrs[prod_ind])):
            attrs_dicts[prod_ind][attr_names[i]] = new_attrs[prod_ind][i]
    
    return attrs_dicts

def load_pickle_obj(file):
    import pickle

    with open(file, 'rb') as f:
        return pickle.load(f)


# XLSX modifier functions
def add_empty_product(product_info, wb):
    products_sheet = wb['Products']
    products_sheet.append(['' for i in range(products_sheet.max_column)])
    row_num = products_sheet.max_row

    # Write the new product code
    last_product_id = products_sheet['A' + str(row_num - 1)].value
    curr_product_id = last_product_id + 1
    products_sheet['A' + str(row_num)] = curr_product_id

    print("Insert new product with ID {} in row {} with Model: {}".format( \
        curr_product_id, row_num, product_info[MODEL_INDX]))

def add_attributes(product_info, attribute_info, wb):
    products_sheet = wb['Products']
    attributes_sheet = wb['ProductAttributes']
    row_num = products_sheet.max_row
    attr_row_num = attributes_sheet.max_row + 1

    (categories_to_attribute, __, attributes_dict) = load_pickle_obj('pkl_files/attributes.pkl')
    
    attr_grp = categories_to_attribute[product_info[CATEG_INDX]]
    attributes = attributes_dict[attr_grp]

    last_product_id = products_sheet['A' + str(row_num - 1)].value
    curr_product_id = last_product_id + 1

    # Data Processing
    attribute_info = static_pre_processing(product_info, attribute_info, attr_grp)
    attribute_info = process_attr_data(attribute_info, attr_grp)
    attribute_info = static_post_processing(product_info, attribute_info, attr_grp)
    # pprint(attribute_info['ΥΛΙΚΟ ΔΕΣΙΜΑΤΟΣ'])
    i = 0
    for attr in attributes:
        # If attribute is empty, don't insert the row
        if attr not in attribute_info or attribute_info[attr][0] == "":
            continue
        attributes_sheet.append(['' for j in range(attributes_sheet.max_column)])
        
        attributes_sheet['C' + str(attr_row_num + i)] = attr
        attributes_sheet['A' + str(attr_row_num + i)] = curr_product_id
        attributes_sheet['B' + str(attr_row_num + i)] = attr_grp

        if attr in attribute_info:
            attributes_sheet['D' + str(attr_row_num + i)] = attribute_info[attr][0]
            attributes_sheet['E' + str(attr_row_num + i)] = attribute_info[attr][1]
        else:
            attributes_sheet['D' + str(attr_row_num + i)] = ''
            attributes_sheet['E' + str(attr_row_num + i)] = ''
        i += 1

def add_product_name(product_info, wb):
    products_sheet = wb['Products']
    row_num = products_sheet.max_row

    # Create product name
    product_name = product_info[MANUF_INDX]  + ' ' \
                 + product_info[COLLEC_INDX] + ' '
    if product_info[FAMILY_INDX] != '':
        product_name += product_info[FAMILY_INDX] + ' '
    product_name += product_info[MODEL_INDX]

    # Write product name
    products_sheet['B' + str(row_num)] = product_name
    products_sheet['C' + str(row_num)] = product_name

def add_description(product_info, wb):
    products_sheet = wb['Products']
    row_num = products_sheet.max_row

    # Define gender
    if product_info[GENDER_INDX] == 'mens':
        gender_en = 'mens'
        gender_el = 'ανδρικό'
    elif product_info[GENDER_INDX] == 'womens':
        gender_en = 'womens'
        gender_el = 'γυναικείο'

    # Greek
    descr_el = 'Κομψό ' + gender_el + ' ρολόι από την εταιρία ' \
             + product_info[MANUF_INDX] \
             + ', Ελβετικής κατασκευής, από τη συλλογή ' \
             + product_info[COLLEC_INDX] 
    if product_info[FAMILY_INDX] != '':
        # Family could be empty
        descr_el += ' ' + product_info[FAMILY_INDX]
    descr_el += ', με κωδικό ' + product_info[MODEL_INDX] + '.'

    # English
    descr_en = 'An elegant ' + gender_en + ' watch by ' \
             + product_info[MANUF_INDX] \
             + ', Swiss made, from the collection ' \
             + product_info[COLLEC_INDX] 
    if product_info[FAMILY_INDX] != '':
        # Family could be empty
        descr_en += ' ' + product_info[FAMILY_INDX]
    descr_en += ', with reference ' + product_info[MODEL_INDX] + '.'

    # Write description
    products_sheet['AE' + str(row_num)] = '<p>' + descr_el + '</p>'
    products_sheet['AF' + str(row_num)] = '<p>' + descr_en + '</p>'

    # Write meta description
    products_sheet['AI' + str(row_num)] = descr_el
    products_sheet['AJ' + str(row_num)] = descr_en

def add_SEO(product_info, wb):
    products_sheet = wb['Products']
    row_num = products_sheet.max_row

    # Create SEO
    SEO = product_info[MANUF_INDX].replace(' ', '_') + '-' \
        + product_info[COLLEC_INDX].replace(' ', '_') + '-'
    if product_info[FAMILY_INDX] != '':
        SEO += product_info[FAMILY_INDX].replace(' ', '_') + '-'
    SEO += product_info[MODEL_INDX].replace(' ', '_')

    # Write SEO
    products_sheet['AD' + str(row_num)] = SEO

def add_model(product_info, wb):
    products_sheet = wb['Products']
    row_num = products_sheet.max_row

    # Get model
    model = product_info[MODEL_INDX]

    # Write model
    products_sheet['M' + str(row_num)] = model

def add_meta_title(product_info, wb):
    products_sheet = wb['Products']
    row_num = products_sheet.max_row

    # Define gender
    if product_info[GENDER_INDX] == 'mens':
        gender_en = 'Mens'
        gender_el = 'Ανδρικό'
    elif product_info[GENDER_INDX] == 'womens':
        gender_en = 'Womens'
        gender_el = 'Γυναικείο'

    # Greek
    if product_info[FAMILY_INDX] != '':
        meta_title_el_m = gender_el + ' Ελβετικό ρολόι - ' \
                        + product_info[MANUF_INDX]  + ' ' \
                        + product_info[FAMILY_INDX] + ' ' \
                        + product_info[MODEL_INDX]  + ' | Eurotimer'
        meta_title_el_l = gender_el + ' Ελβετικό ρολόι - ' \
                        + product_info[MANUF_INDX]  + ' ' \
                        + product_info[COLLEC_INDX] + ' ' \
                        + product_info[FAMILY_INDX] + ' ' \
                        + product_info[MODEL_INDX]  + ' | Eurotimer'
    else:
        meta_title_el_m = gender_el + ' Ελβετικό ρολόι - ' \
                        + product_info[MANUF_INDX]  + ' ' \
                        + product_info[COLLEC_INDX] + ' ' \
                        + product_info[MODEL_INDX]  + ' | Eurotimer'
        meta_title_el_l = meta_title_el_m
    meta_title_el_s = gender_el + ' Ελβετικό ρολόι - ' \
                    + product_info[MANUF_INDX] + ' ' \
                    + product_info[MODEL_INDX] + ' | Eurotimer'
    
    # English
    if product_info[FAMILY_INDX] != '':
        meta_title_en_m = gender_en + ' Watch - Swiss Made ' \
                        + product_info[MANUF_INDX]  + ' ' \
                        + product_info[FAMILY_INDX] + ' ' \
                        + product_info[MODEL_INDX]  + ' | Eurotimer'
        meta_title_en_l = gender_en + ' Watch - Swiss Made ' \
                        + product_info[MANUF_INDX]  + ' ' \
                        + product_info[COLLEC_INDX] + ' ' \
                        + product_info[FAMILY_INDX] + ' ' \
                        + product_info[MODEL_INDX]  + ' | Eurotimer'
    else:
        meta_title_en_m = gender_en + ' Watch - Swiss Made ' \
                        + product_info[MANUF_INDX]  + ' ' \
                        + product_info[COLLEC_INDX] + ' ' \
                        + product_info[MODEL_INDX]  + ' | Eurotimer'
        meta_title_en_l = meta_title_en_m
    meta_title_en_s = gender_en + ' Watch - Swiss Made ' \
                    + product_info[MANUF_INDX] + ' ' \
                    + product_info[MODEL_INDX] + ' | Eurotimer'

    # Select meta title with appropriate length
    if len(meta_title_el_l) <= 60:
        meta_title_el = meta_title_el_l
    elif len(meta_title_el_m) <= 60:
        meta_title_el = meta_title_el_m
    elif len(meta_title_el_s) <= 60:
        meta_title_el = meta_title_el_s
    else:
        print('Warning: Greek meta title is longer than 60 characters!')
        meta_title_el = ''

    if len(meta_title_en_l) <= 60:
        meta_title_en = meta_title_en_l
    elif len(meta_title_en_m) <= 60:
        meta_title_en = meta_title_en_m
    elif len(meta_title_en_s) <= 60:
        meta_title_en = meta_title_en_s
    else:
        print('Warning: English meta title is longer than 60 characters!')
        meta_title_en = ''

    # Wrte meta titles
    products_sheet['AG' + str(row_num)] = meta_title_el
    products_sheet['AH' + str(row_num)] = meta_title_en

def add_price(product_info, wb):
    products_sheet = wb['Products']
    row_num = products_sheet.max_row

    # Get price
    price = product_info[PRICE_INDX]

    # Write price
    products_sheet['Q' + str(row_num)] = price

def add_manufacturer(product_info, wb):
    products_sheet = wb['Products']
    row_num = products_sheet.max_row

    # Check manufacturer
    manuf = ''
    for correct_manuf in MANUFACTURER:
        if product_info[MANUF_INDX]== correct_manuf:
            manuf = product_info[MANUF_INDX]
    if manuf == '':
        print('Warning: invalid manufacturer name!')

    # Write manufacturer
    products_sheet['N' + str(row_num)] = manuf

def add_category(product_info, wb):
    products_sheet = wb['Products']
    row_num = products_sheet.max_row
    categories_dict = load_pickle_obj('pkl_files/categories.pkl')

    # Find category number from dictionary
    categ = closest_match(product_info[CATEG_INDX], categories_dict)

    # Also add the parent categories
    broken_category = categ.split(">")
    parent_categ = broken_category[0]
    categ_val = str(categories_dict[parent_categ])
    for i in range(1, len(broken_category)):
        parent_categ = parent_categ + '>' + broken_category[i]
        categ_val += "," + str( categories_dict[parent_categ] )


    # Write category number
    products_sheet['D' + str(row_num)] = categ_val

def add_status(product_info, wd):
    products_sheet = wb['Products']
    row_num = products_sheet.max_row

    not_hidden = True
    try:
        not_hidden = not product_info[HIDDN_INDX]
    except IndexError:
        not_hidden = True

    if not_hidden:
        products_sheet['AB' + str(row_num)] = 'true'
    else:
        products_sheet['AB' + str(row_num)] = 'false'

def add_image(product_info, wb):
    products_sheet = wb['Products']
    row_num = products_sheet.max_row
    alph = 'abcdefghijklmnopqrstuvwxyz'
    
    # Define image directory
    model = product_info[MODEL_INDX]
    categ_dir = product_info[CATEG_INDX].replace('>', '/')
    image_dir = 'catalog/product/' + categ_dir + '/' + model + '.jpg'

    # Write image directory
    if int('0' + str(product_info[IMG_INDX])) == 0:
        print(product_info[IMG_INDX] + ' is 0')
        products_sheet['O' + str(row_num)] = 'catalog/product/placeholder.jpg'
        return

    if int(product_info[IMG_INDX]) == 1:
        products_sheet['O' + str(row_num)] = image_dir
        return

    products_sheet['O' + str(row_num)] = image_dir[:-4] + 'a.jpg'

    # Extra images
    sheet = wb["AdditionalImages"]

    for i in range(1, int(product_info[IMG_INDX])):
        sheet.append(['' for i in range(sheet.max_column)])
        row = sheet.max_row
        last_product_id = products_sheet['A' + str(row_num - 1)].value
        curr_product_id = last_product_id + 1
        image_dir = image_dir[:-5] + alph[i] + ".jpg"
        sheet['A' + str(row)] = curr_product_id
        sheet['B' + str(row)] = image_dir
        sheet['C' + str(row)] = i

def add_discount(product_info, wb):
    discount = int('0' + str(product_info[DISC_INDX]).replace('-', ''))
    if discount == 0: return
    price    = int(product_info[PRICE_INDX])

    products_sheet = wb['Products']
    discounts_sheet = wb['Specials']
    row = discounts_sheet.max_row + 1
    row_num = products_sheet.max_row
    last_product_id = products_sheet['A' + str(row_num - 1)].value
    curr_product_id = last_product_id + 1
    discounts_sheet.append(['' for i in range(discounts_sheet.max_column)])

    discounts_sheet['A' + str(row)] = curr_product_id
    discounts_sheet['B' + str(row)] = 'Default'
    discounts_sheet['C' + str(row)] = '0'
    discounts_sheet['D' + str(row)] = price*(1 - discount/100)
    discounts_sheet['E' + str(row)] = '0000-00-00'
    discounts_sheet['F' + str(row)] = '0000-00-00'

def add_misc(product_info, wb):
    products_sheet = wb['Products']
    row_num = products_sheet.max_row

    products_sheet['L'  + str(row_num)] = product_info[QUANT_INDX]      #quantity
    products_sheet['P'  + str(row_num)] = 'yes'  #shipping
    products_sheet['R'  + str(row_num)] = 0      #points
    products_sheet['S'  + str(row_num)] = '2018-07-26 14:00:00' #date_added
    products_sheet['T'  + str(row_num)] = '2018-07-26 14:00:00'#date_modified
    products_sheet['U'  + str(row_num)] = '2018-07-26' #date_available
    products_sheet['V'  + str(row_num)] = 0      #weight
    products_sheet['W'  + str(row_num)] = 'kg'   #weight_unit
    products_sheet['X'  + str(row_num)] = 0      #length
    products_sheet['Y'  + str(row_num)] = 0      #width
    products_sheet['Z'  + str(row_num)] = 0      #height
    products_sheet['AA' + str(row_num)] = 'cm'   #length_unit
    products_sheet['AC' + str(row_num)] = 0      #tax_class_id
    products_sheet['AM' + str(row_num)] = 6      #stock_status_id
    products_sheet['AN' + str(row_num)] = 0      #store_ids
    products_sheet['AS' + str(row_num)] = 1      #sort_order
    products_sheet['AT' + str(row_num)] = 'true' #subtract
    products_sheet['AU' + str(row_num)] = 1      #minimum


if __name__ == '__main__':

    if len(argv) < 4:
        print('arg1: specs.tsv | arg2: attrs.tsv | arg3: products.xlsx')
        exit(1)

    SPECS_TSV = argv[1]
    ATTRS_TSV = argv[2]
    products_xlsx = argv[3]

    wb = load_workbook(products_xlsx)
    new_products = open_new_products(SPECS_TSV)
    new_attributes = open_product_attributes(ATTRS_TSV)
    products = 0
    
    # Iterate the inputs
    for product, attributes in zip(new_products, new_attributes):
        add_empty_product(product, wb)
        add_attributes(product, attributes, wb)
        add_product_name(product, wb)
        add_description(product, wb)
        add_SEO(product, wb)
        add_model(product, wb)
        add_meta_title(product, wb)
        add_price(product, wb)
        add_manufacturer(product, wb)
        add_category(product, wb)
        add_image(product, wb)
        add_discount(product, wb)
        add_status(product, wb)
        add_misc(product, wb)
        products += 1

    # Cleanup and Save to file
    print("Added {} products.".format(products))
    cleanup(wb)
    wb.save(products_xlsx)
