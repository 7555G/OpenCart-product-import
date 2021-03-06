#!/usr/bin/env python
# -*- coding: utf-8 -*-

from sys import argv
from pprint import pprint
from openpyxl import Workbook, load_workbook

from libs.transformations import COLUMN_TRANSF_RULES
from libs.color import get_color
from libs.utilities import *

# Maurice Laxroix Strap|Bracelet Color Material Info Number : Google
# Maurice Laxroix Strap|Bracelet Color Material Buckle Info Number : Site 
MANUFACTURER = "Jos Von Arx"
CATEGORY = "STRAPS>MAURICE LACROIX"
OFFER_CATEG = "OFFERS>ACCESSORIES"

def cleanup(wb):
    products_sheet = wb['Products']
    products_sheet.append(['' for i in range(products_sheet.max_column)])
    row_num = products_sheet.max_row

    # print("Row to clean is: ", row_num)
    products_sheet.delete_rows(row_num)

def open_product_attributes(input_file):
    # The tsv will be converted to a list of dictionaries
    # Now we index as [prod_num]["type"]
    new_attrs = open(input_file, encoding = 'utf-8')
    new_attrs = [line.replace('\n','') for line in new_attrs] 
    new_attrs = [line.split('\t') for line in new_attrs]
    attr_names = new_attrs.pop(0)

    attrs_dicts = [{} for prod in new_attrs]
    for prod_ind in range(len(new_attrs)):
        for i in range(len(new_attrs[prod_ind])):
            attrs_dicts[prod_ind][attr_names[i]] = new_attrs[prod_ind][i]
    
    return attrs_dicts

def load_pickle_obj(file):
    import pickle

    with open(file, 'rb') as f:
        return pickle.load(f)

# XLSX modifier functions
def add_empty_product(product_info, wb):
    products_sheet = wb['Products']
    products_sheet.append(['' for i in range(products_sheet.max_column)])
    row_num = products_sheet.max_row

    # Write the new product code
    last_product_id = products_sheet['A' + str(row_num - 1)].value
    curr_product_id = last_product_id + 1
    products_sheet['A' + str(row_num)] = curr_product_id
    
    print("Insert new product with ID {} in row {} with Model: {}".format( \
        curr_product_id, row_num, product_info["number"]))

def process_attr_data(product, attr_grp):
    for attr in COLUMN_TRANSF_RULES[attr_grp]:
        # First create the list(of lists) based on the priorities
        transformations_ordered = [[] for i in range(len(COLUMN_TRANSF_RULES[attr_grp][attr]) + 1000)] 
        
        for word, transf in COLUMN_TRANSF_RULES[attr_grp][attr].items():
            transformations_ordered[transf[-1]].append(word)
            # [ ['STEEL','S.S'], ['Ceramic'], ... ]
            
        # Clean empty lists
        transformations_ordered = [lst for lst in transformations_ordered if lst]
        # Watch out for no data
        try:
            if not product[attr]:
                product[attr] = ""
                continue
        except:
            product[attr] = ""


        new_val_el = ""
        new_val_en = ""
        for transf in transformations_ordered:
            possible_words = [word for word in transf]

            match_word = ""
            # Now check if any of the words are in the text
            for word in possible_words:
                if word == "__*":
                    match_word = word
                    break

                if "__color" in word:
                    colors_vals = get_color(word.lower(), replace_chars("/|", product[attr].lower(), " "))
                    if colors_vals:
                        match_word = word
                        break

                elif word.upper() in product[attr].upper():
                    match_word = word
                    break

            # Match with the first word from the list
            if match_word:
                new_val_en += ' ' + COLUMN_TRANSF_RULES[attr_grp][attr][match_word][0]
                new_val_el += ' ' + COLUMN_TRANSF_RULES[attr_grp][attr][match_word][1]

                # If the pattern matched we have the colors that need to be replaced
                if "__color" in match_word:
                    colors_vals = get_color(match_word.lower(), replace_chars("/|", product[attr].lower(), " "))
                    new_val_el = new_val_el.replace("__color", colors_vals[1])
                    new_val_en = new_val_en.replace("__color", colors_vals[0])
        
        if not new_val_el:
            product[attr] = [product[attr], product[attr]]
        elif new_val_el == " __clear":
            # Special commands
            product[attr] = ["", ""]
        else:
            # Remove the first space
            new_val_el = new_val_el[1:][0].upper() + new_val_el[2:]
            new_val_en = new_val_en[1:][0].upper() + new_val_en[2:]
            if new_val_el[-1] == ',':
                new_val_el = new_val_el[:-1]
                new_val_en = new_val_en[:-1]
            product[attr] = [new_val_el, new_val_en]

    # Lastly all attributes that weren't filtered, convert them to [greek, engl] format
    for attr in product:
        if not isinstance(product[attr], list) and attr in COLUMN_TRANSF_RULES[attr_grp]:
            product[attr] = [product[attr], product[attr]]

    return product        

def static_pre_processing(product_info, attr_grp):
    # Apply rules to the data before inserting it
    product_info['name_material'] = product_info['ΥΛΙΚΟ']

    return product_info

def static_post_processing(product_info, attr_grp):  
    product_info['ΠΛΗΡΟΦΟΡΙΕΣ'] = ['', '']
    if product_info['info gr']:
        product_info['ΠΛΗΡΟΦΟΡΙΕΣ'] = [product_info['info gr'], product_info['info en']]

    product_info['ΥΛΙΚΟ'][0] = product_info['ΥΛΙΚΟ'][0].replace(',', ' και')
    product_info['ΥΛΙΚΟ'][1] = product_info['ΥΛΙΚΟ'][1].replace(',', ' and')

    product_info['ΔΙΑΣΤΑΣΕΙΣ'] = [product_info['ΔΙΑΣΤΑΣΕΙΣ'], product_info['ΔΙΑΣΤΑΣΕΙΣ']]

    product_info['brand'] = product_info['brand'].title()
    product_info['category_short'] = product_info['category'].split('>')[-1].lower()[:-1]

    return product_info

def add_attributes(product_info, wb):
    products_sheet = wb['Products']
    attributes_sheet = wb['ProductAttributes']
    row_num = products_sheet.max_row
    attr_row_num = attributes_sheet.max_row + 1

    (categories_to_attribute, __, attributes_dict) = load_pickle_obj('pkl_files/attributes.pkl')

    attr_grp = categories_to_attribute[product_info['category']]
    attributes = attributes_dict[attr_grp]

    last_product_id = products_sheet['A' + str(row_num - 1)].value
    curr_product_id = last_product_id + 1

    # Data Processing
    attribute_info = static_pre_processing(product_info, attr_grp)
    attribute_info = process_attr_data(attribute_info, attr_grp)
    attribute_info = static_post_processing(attribute_info, attr_grp)

    i = 0
    for attr in attributes:
        # If attribute is empty, don't insert the row
        if attr not in attribute_info or attribute_info[attr][0] == "":
            continue
        attributes_sheet.append(['' for j in range(attributes_sheet.max_column)])
        
        attributes_sheet['C' + str(attr_row_num + i)] = attr
        attributes_sheet['A' + str(attr_row_num + i)] = curr_product_id
        attributes_sheet['B' + str(attr_row_num + i)] = attr_grp

        if attr in attribute_info:
            attributes_sheet['D' + str(attr_row_num + i)] = attribute_info[attr][0]
            attributes_sheet['E' + str(attr_row_num + i)] = attribute_info[attr][1]
        else:
            attributes_sheet['D' + str(attr_row_num + i)] = ''
            attributes_sheet['E' + str(attr_row_num + i)] = ''
        i += 1

def add_product_name_and_meta_title(product_info, wb):
    products_sheet = wb['Products']
    row_num = products_sheet.max_row

    # check if has no reference number
    if product_info['number'][0] == '-':
        model = ''
    else:
        model = ', ' + product_info["number"]

    # if made of leather, don't mention brown color
    if 'leather' in product_info['ΥΛΙΚΟ'][1].lower() and \
       'brown'   in product_info['ΧΡΩΜΑ'][1].lower():
        color = ['', '']
    else:
        color = [product_info['ΧΡΩΜΑ'][0] + ' ',
                 product_info['ΧΡΩΜΑ'][1] + ' ']

    # Create product name
    if "bracelet" in product_info['category_short']:
        title_el = product_info['brand'] + ' ' + product_info['gender'][0] \
                 + ' Βραχιόλι από ' + color[0].title() \
                 + product_info['ΥΛΙΚΟ'][0].title() + model
        title_en = product_info['brand'] + ' ' + product_info['gender'][1] \
                 + ' Bracelet made of ' + color[1].title() \
                 + product_info['ΥΛΙΚΟ'][1].title() + model

    if "pen" in product_info['category_short']:
        title_el = product_info['brand'] + ' Στυλό από ' \
                 + product_info['name_material'][0].title() + model
        title_en = product_info['brand'] + ' Pen made of ' \
                 + product_info['name_material'][1].title() + model

    if "wallet" in product_info['category_short']:
        title_el = product_info['brand'] + ' ' + product_info['gender'][0] + ' ' \
                 + color[0].title() + 'Δερμάτινο Πορτοφόλι' + model
        title_en = product_info['brand'] + ' ' + product_info['gender'][1] + ' ' \
                 + color[1].title() + 'Leather Wallet' + model

    if "card holder" in product_info['category_short']:
        if color[0] != '':
            if color[0][-2] == 'ο':
                color[0] = color[0][:-2] + 'η '
        title_el = product_info['brand'] + ' ' + color[0].title() \
                 + 'Δερμάτινη Θήκη για Κάρτες' + model 
        title_en = product_info['brand'] + ' ' + color[1].title() \
                 + 'Leather Card Holder' + model

    if "pendant" in product_info['category_short']:
        if 'cross' in product_info['info en'][1].lower():
            temp_categ = [' ' + product_info['gender'][0] + 'ς Σταυρός', \
                          ' ' + product_info['gender'][1] + ' Cross']
        else:
            temp_categ = [' ' + product_info['gender'][0] + ' Κολιέ', \
                          ' ' + product_info['gender'][1] + ' Pendant']
        title_el = product_info['brand'] + temp_categ[0] + ' από ' \
                 + product_info['ΥΛΙΚΟ'][0].title() + model
        title_en = product_info['brand'] + temp_categ[1] + ' made of ' \
                 + product_info['ΥΛΙΚΟ'][1].title() + model

    if "cufflink" in product_info['category_short']:
        title_el = product_info['brand'] + ' Μανικετόκουμπα' + model
        title_en = product_info['brand'] + ' Cufflinks' + model

    if "gift set" in product_info['category_short']:
        title_el = product_info['brand'] + ' Σετ Δώρων από ' \
                 + product_info['name_material'][0].title() + model
        title_en = product_info['brand'] + ' Gift Set made of ' \
                 + product_info['name_material'][1].title() + model
        if 'leather' in product_info['name_material'][1].lower():
            title_el = product_info['brand'] + ' ' \
                     + product_info['name_material'][0].title() \
                     + ' Σετ Δώρων' + model
            title_en = product_info['brand'] + ' ' \
                     + product_info['name_material'][1].title() \
                     + ' Gift Set' + model

    if "key holder" in product_info['category_short']:
        title_el = product_info['brand'] + ' Κλειδοθήκη από ' \
                 + product_info['ΥΛΙΚΟ'][0].title() + model
        title_en = product_info['brand'] + ' Key holder made of ' \
                 + product_info['ΥΛΙΚΟ'][1].title() + model

    if "money clip" in product_info['category_short']:
        title_el = product_info['brand'] + ' Χρηματοπιάστρα από ' \
                 + product_info['ΥΛΙΚΟ'][0].title() + model
        title_en = product_info['brand'] + ' Money Clip made of ' \
                 + product_info['ΥΛΙΚΟ'][1].title() + model

    if "tiebar" in product_info['category_short']:
        title_el = product_info['brand'] + ' Γραβατοπιάστρα από ' \
                 + product_info['name_material'][0].title() + model
        title_en = product_info['brand'] + ' Tiebar made of ' \
                 + product_info['name_material'][1].title() + model

    meta_title_el = title_el + ' | Eurotimer'
    meta_title_en = title_en + ' | Eurotimer'

    # Write product name
    products_sheet['B' + str(row_num)] = title_el
    products_sheet['C' + str(row_num)] = title_en

    # Write meta titles
    products_sheet['AG' + str(row_num)] = meta_title_el
    products_sheet['AH' + str(row_num)] = meta_title_en

def add_description(product_info, wb):
    products_sheet = wb['Products']
    row_num = products_sheet.max_row

    # if made of leather, don't mention brown color
    if 'leather' in product_info['ΥΛΙΚΟ'][1].lower() and \
       'brown'   in product_info['ΧΡΩΜΑ'][1].lower():
        color = ['', '']
    else:
        color = [product_info['ΧΡΩΜΑ'][0] + ' ',
                 product_info['ΧΡΩΜΑ'][1] + ' ']

    if "bracelet" in product_info['category_short']:
        descr_el = product_info['gender'][0] + ' βραχιόλι από ' + color[0].lower() \
                 + product_info['ΥΛΙΚΟ'][0].lower() + ', της εταιρίας ' \
                 + product_info['brand'] + '. '
        descr_en = product_info['gender'][1] + ' bracelet made of ' + color[1].lower() \
                 + product_info['ΥΛΙΚΟ'][1].lower() + ', by ' \
                 + product_info['brand'] + '. '

    if "pen" in product_info['category_short']:
        descr_el = 'Στυλό από ' + product_info['name_material'][0].lower() \
                 + ', της εταιρίας ' + product_info['brand'] + '. '
        descr_en = 'Pen made of ' + product_info['name_material'][1].lower() \
                 + ', by ' + product_info['brand'] + '. '

    if "wallet" in product_info['category_short']:
        descr_el = product_info['gender'][0] + ' πορτοφόλι από ' \
                 + product_info['ΥΛΙΚΟ'][0].lower() + ', της εταιρίας ' \
                 + product_info['brand'] + '. '
        descr_en = product_info['gender'][1] + ' wallet made of ' \
                 + product_info['ΥΛΙΚΟ'][1].lower() + ', by ' \
                 + product_info['brand'] + '. '

    if "card holder" in product_info['category_short']:
        descr_el = 'Θήκη για κάρτες από ' + product_info['ΥΛΙΚΟ'][0].lower() \
                 + ', της εταιρίας ' + product_info['brand'] + '. '
        descr_en = 'Card holder made of ' + product_info['ΥΛΙΚΟ'][1].lower() \
                 + ', by ' + product_info['brand'] + '. '

    if "pendant" in product_info['category_short']:
        if 'cross' in product_info['info en'][1].lower():
            temp_categ = [' ' + product_info['gender'][0] + 'ς Σταυρός', \
                          ' ' + product_info['gender'][1] + ' Cross']
        else:
            temp_categ = [' ' + product_info['gender'][0] + ' Κολιέ', \
                          ' ' + product_info['gender'][1] + ' Pendant']
        descr_el = temp_categ[0].capitalize() + ' από ' + product_info['ΥΛΙΚΟ'][0].lower() \
                 + ', της εταιρίας ' + product_info['brand'] + '. '
        descr_en = temp_categ[1].capitalize() + ' από ' + product_info['ΥΛΙΚΟ'][1].lower() \
                 + ', της εταιρίας ' + product_info['brand'] + '. '

    if "cufflink" in product_info['category_short']:
        descr_el = 'Μανικετόκουμπα από ' + product_info['name_material'][0].lower() \
                 + ', από την εταιρία ' + product_info['brand'] + '. '
        descr_en = 'Cufflinks made of ' + product_info['name_material'][1].lower() \
                 + ', by ' + product_info['brand'] + '. '

    if "gift set" in product_info['category_short']:
        descr_el = 'Σετ δώρων από ' + product_info['name_material'][0].lower() \
                 + ', της εταιρίας ' + product_info['brand'] + '. '
        descr_en = 'Gift set made of ' + product_info['name_material'][1].lower() \
                 + ', by ' + product_info['brand'] + '. '
        if 'leather' in product_info['name_material'][1].lower():
            descr_el = color[0].capitalize() + product_info['name_material'][0].lower() \
                     + ' σετ δώρων, από την εταιρία ' + product_info['brand'] + '. '
            descr_en = color[1].capitalize() + product_info['name_material'][1].lower() \
                     + 'gift set, by ' + product_info['brand'] + '. '

    if "key holder" in product_info['category_short']:
        descr_el = product_info['brand'] + ' Κλειδοθήκη από ' \
                 + product_info['ΥΛΙΚΟ'][0].lower()
        descr_en = product_info['brand'] + ' Key holder made of ' \
                 + product_info['ΥΛΙΚΟ'][1].lower()

    if "money clip" in product_info['category_short']:
        descr_el = product_info['brand'] + ' Χρηματοπιάστρα από ' \
                 + product_info['ΥΛΙΚΟ'][0].lower()
        descr_en = product_info['brand'] + ' Money Clip made of ' \
                 + product_info['ΥΛΙΚΟ'][1].lower()

    if "tiebar" in product_info['category_short']:
        descr_el = product_info['brand'] + ' Γραβατοπιάστρα από ' \
                 + product_info['name_material'][0].lower()
        descr_en = product_info['brand'] + ' Tiebar made of ' \
                 + product_info['name_material'][1].lower()

    if product_info['info gr']:
        descr_el += product_info['info gr']
        descr_en += product_info['info en']

    meta_descr_el = descr_el
    meta_descr_en = descr_en

    # Write description
    products_sheet['AE' + str(row_num)] = '<p>' + descr_el + '</p>'
    products_sheet['AF' + str(row_num)] = '<p>' + descr_en + '</p>'

    # Write meta description
    products_sheet['AI' + str(row_num)] = meta_descr_el
    products_sheet['AJ' + str(row_num)] = meta_descr_en

def add_SEO(product_info, wb):
    products_sheet = wb['Products']
    row_num = products_sheet.max_row

    SEO = (product_info['brand'] + '-' + product_info['category_short'] + '-' \
        + product_info['number']).replace(' ', '_')

    # Write SEO
    products_sheet['AD' + str(row_num)] = SEO

def add_model(product_info, wb):
    products_sheet = wb['Products']
    row_num = products_sheet.max_row

    # Get model
    if product_info['number'][0] == '-':
        model = '-'
    else:
        model = product_info["number"]

    # Write model
    products_sheet['M' + str(row_num)] = model

def add_price(product_info, wb):
    products_sheet = wb['Products']
    row_num = products_sheet.max_row

    # Get price
    price = product_info["price"]

    # Write price
    products_sheet['Q' + str(row_num)] = price

def add_manufacturer(product_info, wb):
    products_sheet = wb['Products']
    row_num = products_sheet.max_row

    # Check manufacturer
    manuf = product_info['brand']
    if manuf == '':
        print('Warning: invalid manufacturer name!')

    # Write manufacturer
    products_sheet['N' + str(row_num)] = manuf

def add_category(product_info, wb):
    products_sheet = wb['Products']
    row_num = products_sheet.max_row
    categories_dict = load_pickle_obj('pkl_files/categories.pkl')

    # Find category number from dictionary
    categ = closest_match(product_info['category'], categories_dict)

    # Also add the parent categories
    broken_category = categ.split(">")
    parent_categ = broken_category[0]
    categ_val = str(categories_dict[parent_categ])

    for i in range(1, len(broken_category)):
        parent_categ = parent_categ + '>' + broken_category[i]
        categ_val += "," + str( categories_dict[parent_categ] )

    # Write category number
    discount = int('0' + str(product_info['discount']).replace('-', ''))
    if discount != 0:
        categ_val += "," + str( categories_dict["OFFERS"])
        categ_val += "," + str( categories_dict[OFFER_CATEG])

    products_sheet['D' + str(row_num)] = categ_val

def add_status(product_info, wb):
    products_sheet = wb['Products']
    row_num = products_sheet.max_row

    if not product_info['hidden']:
        products_sheet['AB' + str(row_num)] = 'true'
    else:
        products_sheet['AB' + str(row_num)] = 'false'

def add_image(product_info, wb):
    products_sheet = wb['Products']
    row_num = products_sheet.max_row
    alph = 'abcdefghijklmnopqrstuvwxyz'

    # Write image directory
    if int('0' + str(product_info['img count'])) == 0:
        print(product_info['img count'] + ' is 0')
        image_dir = 'catalog/product/placeholder.jpg'
    elif int(product_info['img count']) == 1:
        image_dir = 'catalog/product/' + product_info['category'].replace('>', '/') \
                  + '/' + product_info["number"] + ".jpg"
    else:
        image_dir = 'catalog/product/' + product_info['category'].replace('>', '/') \
                  + '/' + product_info["number"] + "a.jpg"

    products_sheet['O' + str(row_num)] = image_dir

    # Extra images
    sheet = wb["AdditionalImages"]
    last_product_id = products_sheet['A' + str(row_num - 1)].value
    curr_product_id = last_product_id + 1
    img_count = int(product_info['img count'])

    for i in range(1, img_count):
        sheet.append(['' for i in range(sheet.max_column)])
        row = sheet.max_row
        image_dir = image_dir[:-5] + alph[i] + ".jpg"
        sheet['A' + str(row)] = curr_product_id
        sheet['B' + str(row)] = image_dir
        sheet['C' + str(row)] = i

    # Add category-specific images
    if "bracelet" in product_info['category_short']:
        sheet.append(['' for i in range(sheet.max_column)])
        row = sheet.max_row
        image_dir = 'catalog/product/' + product_info['category'].replace('>', '/') \
                  + '/' + 'bracelet-box.jpg'
        sheet['A' + str(row)] = curr_product_id
        sheet['B' + str(row)] = image_dir
        sheet['C' + str(row)] = 1

    if "cufflink" in product_info['category_short']:
        sheet.append(['' for i in range(sheet.max_column)])
        row = sheet.max_row
        image_dir = 'catalog/product/' + product_info['category'].replace('>', '/') \
                  + '/' + 'cufflinks-box.jpg'
        sheet['A' + str(row)] = curr_product_id
        sheet['B' + str(row)] = image_dir
        sheet['C' + str(row)] = 1

    if product_info['category_short'] == "pen":
        sheet.append(['' for i in range(sheet.max_column)])
        row = sheet.max_row
        image_dir = 'catalog/product/' + product_info['category'].replace('>', '/') \
                  + '/' + 'pen-box-open.jpg'
        sheet['A' + str(row)] = curr_product_id
        sheet['B' + str(row)] = image_dir
        sheet['C' + str(row)] = 1

        sheet.append(['' for i in range(sheet.max_column)])
        row = sheet.max_row
        image_dir = 'catalog/product/' + product_info['category'].replace('>', '/') \
                  + '/' + 'pen-box.jpg'
        sheet['A' + str(row)] = curr_product_id
        sheet['B' + str(row)] = image_dir
        sheet['C' + str(row)] = 2

def add_discount(product_info, wb):
    discount = int('0' + str(product_info['discount']).replace('-', ''))
    if discount == 0: return
    price    = int(product_info['price'])

    products_sheet = wb['Products']
    discounts_sheet = wb['Specials']
    row = discounts_sheet.max_row + 1
    row_num = products_sheet.max_row
    last_product_id = products_sheet['A' + str(row_num - 1)].value
    curr_product_id = last_product_id + 1
    discounts_sheet.append(['' for i in range(discounts_sheet.max_column)])

    discounts_sheet['A' + str(row)] = curr_product_id
    discounts_sheet['B' + str(row)] = 'Default'
    discounts_sheet['C' + str(row)] = '0'
    discounts_sheet['D' + str(row)] = price*(1 - discount/100)
    discounts_sheet['E' + str(row)] = '0000-00-00'
    discounts_sheet['F' + str(row)] = '0000-00-00'

def add_filters(product_info, wb):
    products_sheet = wb['Products']
    filters_sheet = wb['ProductFilters']
    row = filters_sheet.max_row + 1
    row_num = products_sheet.max_row
    last_product_id = products_sheet['A' + str(row_num - 1)].value
    curr_product_id = last_product_id + 1

    categ = product_info["category"].split(">")[0]
    if categ == "ACCESSORIES":
        filters_sheet.append(['' for i in range(filters_sheet.max_column)])
        filters_sheet['A' + str(row)] = curr_product_id
        filters_sheet['B' + str(row)] = "Φύλο"
        filters_sheet['C' + str(row)] = product_info['gender'][0]
    else:
        if "steel" in product_info['ΥΛΙΚΟ'][1].lower():
            filters_sheet.append(['' for i in range(filters_sheet.max_column)])
            filters_sheet['A' + str(row)] = curr_product_id
            filters_sheet['B' + str(row)] = "Μέταλλο"
            filters_sheet['C' + str(row)] = "Ατσάλι"
            row += 1
        elif "brass" in product_info['ΥΛΙΚΟ'][1].lower():
            filters_sheet.append(['' for i in range(filters_sheet.max_column)])
            filters_sheet['A' + str(row)] = curr_product_id
            filters_sheet['B' + str(row)] = "Μέταλλο"
            filters_sheet['C' + str(row)] = "Ορείχαλκος"
            row += 1 
        else:
            pprint("Didn't expect that material")
            exit()

        filters_sheet.append(['' for i in range(filters_sheet.max_column)])
        filters_sheet['A' + str(row)] = curr_product_id
        filters_sheet['B' + str(row)] = "Μάρκα"
        filters_sheet['C' + str(row)] = "Jos Von Arx"

def add_misc(product_info, wb):
    products_sheet = wb['Products']
    row_num = products_sheet.max_row
    
    products_sheet['L'  + str(row_num)] = product_info["stock"]      #quantity
    products_sheet['P'  + str(row_num)] = 'yes'  #shipping
    products_sheet['R'  + str(row_num)] = 0      #points
    products_sheet['S'  + str(row_num)] = '2018-09-25 18:00:00' #date_added
    products_sheet['T'  + str(row_num)] = '2018-09-25 18:00:00'#date_modified
    products_sheet['U'  + str(row_num)] = '2018-09-25' #date_available
    products_sheet['V'  + str(row_num)] = 0      #weight
    products_sheet['W'  + str(row_num)] = 'kg'   #weight_unit
    products_sheet['X'  + str(row_num)] = 0      #length
    products_sheet['Y'  + str(row_num)] = 0      #width
    products_sheet['Z'  + str(row_num)] = 0      #height
    products_sheet['AA' + str(row_num)] = 'cm'   #length_unit
    products_sheet['AC' + str(row_num)] = 0      #tax_class_id
    products_sheet['AM' + str(row_num)] = 6      #stock_status_id
    products_sheet['AN' + str(row_num)] = 0      #store_ids
    products_sheet['AS' + str(row_num)] = 1      #sort_order
    products_sheet['AT' + str(row_num)] = 'true' #subtract
    products_sheet['AU' + str(row_num)] = 1      #minimum


if __name__ == '__main__':

    if len(argv) < 3:
        print('arg1: specs.tsv | arg2: products.xlsx')
        exit(1)

    SPECS_TSV = argv[1]
    products_xlsx = argv[2]

    wb = load_workbook(products_xlsx)
    new_products = open_product_attributes(SPECS_TSV)
    products_count = 0
    

    # Iterate the inputs
    for product in new_products:
        if not product["category"]: continue
        add_empty_product(product, wb)
        add_attributes(product, wb)
        add_product_name_and_meta_title(product, wb)
        add_description(product, wb)
        add_SEO(product, wb)
        add_model(product, wb)
        add_price(product, wb)
        add_manufacturer(product, wb)
        add_category(product, wb)
        add_status(product, wb)
        add_image(product, wb)
        add_discount(product, wb)
        add_filters(product, wb)
        add_misc(product, wb)
        products_count += 1

    # Cleanup and Save to file
    print("Added {} products.".format(products_count))
    cleanup(wb)
    wb.save(products_xlsx)
