#!/usr/bin/env python

from sys import argv
from pprint import pprint
from openpyxl import Workbook, load_workbook

# Global Data
products_xlsx = ""
categories_dict = {}

MODEL_INDX = 0
MANUF_INDX = 1
CATEG_INDX = 2
COLLEC_INDX = 3
FAMILY_INDX = 4
GENDER_INDX = 5


def open_new_products(input_file):
    new_products = open(input_file)
    new_products = [line.replace("\n","") for line in new_products]

    # Remove commas and trim spaces
    new_products = [line.split(",") for line in new_products]

    for product in range(len(new_products)):
        for entry in range(len(new_products[product])):
            new_products[product][entry] = new_products[product][entry].strip()

    # Handle the Category
    for product in range(len(new_products)):
        new_products[product][2] = new_products[product][2].replace(" ", "")
        new_products[product][2] = new_products[product][2].upper()
    
    return new_products

def load_categories():
    import pickle

    with open("categories.pkl", "rb") as f:
        return pickle.load(f)

# XLSX modifier funcions
def add_empty_product(wb):
    products_sheet = wb.active

    products_sheet.append(["" for i in range(products_sheet.max_column)])
    row_num = products_sheet.max_row
   
    # Edit the product code
    print(row_num)
    last_product_id = products_sheet['A' + str(row_num - 1)].value
    curr_product_id = last_product_id + 1
    products_sheet['A' + str(row_num)] = curr_product_id

    wb.save(products_xlsx)

def add_product_descrition(product_info, wb):
    products_sheet = wb.active

    # Edit the product code
    row_num = products_sheet.max_row
    curr_product_id = products_sheet['A' + str(row_num)].value

    gender_en = "mens"
    gender_el = "ανδρικό"
    if product_info[GENDER_INDX] == "female":
        gender_en = "womens"
        gender_el = "γυναικείο"

    # Ελληνικά
    descr_el = "<p>Κομψό " + gender_el + " ρολόι της εταιρίας " + product_info[MANUF_INDX] + ", "
    descr_el += "Ελβετικής κασκευής, από τη συλλογή " + product_info[COLLEC_INDX] 
    if product_info[FAMILY_INDX] != "":
        # Family could be empty
        descr_el += " " + product_info[FAMILY_INDX]
    descr_el += ", με κωδικό " + product_info[MODEL_INDX] + ".</p>"

    # English
    descr_en = "<p>Elegant " + gender_en + " watch by " + product_info[MANUF_INDX] + ", Swiss made, "
    descr_en += "from the collection " + product_info[COLLEC_INDX] 
    if product_info[FAMILY_INDX] != "":
        # Family could be empty
        descr_en+= " " + product_info[FAMILY_INDX]
    descr_en += ", with reference " + product_info[MODEL_INDX] + ".</p>"

    # Save
    products_sheet['AE' + str(row_num)] = descr_el
    products_sheet['AF' + str(row_num)] = descr_en
    wb.save(products_xlsx)


if __name__ == "__main__":

    if len(argv) < 3:
        print("arg1: input_file, arg2: products.xlsx, arg3: images_dir")
        exit(1)

    input_file = argv[1]
    products_xlsx = argv[2]
    images_dir = argv[3]

    wb = load_workbook(products_xlsx)
    new_products = open_new_products(input_file)
    categories_dict = load_categories()
    pprint(new_products)

    # Iterate the inputs
    for product in new_products:
        add_empty_product(wb)
        add_product_descrition(product, wb)
