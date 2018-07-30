#!/usr/bin/env python

from sys import argv
from pprint import pprint
from openpyxl import Workbook, load_workbook
import csv


# Global variables
spec_order = {'model'       : 0,
              'manufacturer': 1,
              'category'    : 2,
              'collection'  : 3,
              'family'      : 4,
              'gender'      : 5,
              'price'       : 6}


def open_specs_file(file_path):
    spec_list = open(file_path)
    spec_list = [line.replace('\n','') for line in spec_list]

    # Remove commas and trim spaces
    spec_list = [line.split(',') for line in spec_list]

    for product in range(len(spec_list)):
        for entry in range(len(spec_list[product])):
            spec_list[product][entry] = spec_list[product][entry].strip() 

    # Create specs dict
    spec = {}
    for i in range(len(spec_list[0])):
        spec[spec_list[0][i]] = spec_list[1][i]

    return spec


def write_data_file(file_path, data):
    file = open(file_path)


def is_column(string):
    if string[0] != '-':
        return 1
    else:
        return 0

def get_data(stock_sheet, product_data, row, spec_dict):
    for item in spec_dict:
        if is_column(spec_dict[item]):
            product_data[row][spec_order[item]] = \
                         stock_sheet[spec_dict[item] + str(row)].value
        else:
            product_data[row][spec_order[item]] = spec_dict[item][1:]


if __name__ == '__main__':

    if len(argv) != 4:
        print('arg1: stock.xlsx, arg2: data.in, arg3: specs.txt')
        exit(1)

    stock_xlsx = argv[1]
    data_in    = argv[2]
    specs_file = argv[3]

    spec = open_specs_file(specs_file)
    wb = load_workbook(stock_xlsx, data_only=True)
    stock_sheet = wb.active
    rows = stock_sheet.max_row
    product_data =[[0 for x in range(len(spec))] \
                      for y in range(rows)] 

    # Collect the data from the spreadsheet to a variable
    for row in range(1, rows):
        get_data(stock_sheet, product_data, row, spec)
    
    # Write data to file
    myfile = open(data_in, 'w')
    writer = csv.writer(myfile, dialect='excel', lineterminator='\n')
    writer.writerows(product_data) 
