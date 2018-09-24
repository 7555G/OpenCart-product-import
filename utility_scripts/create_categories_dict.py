#!/usr/bin/env python

from openpyxl import Workbook, load_workbook
from pprint import pprint
from sys import argv
import pickle

if __name__ == "__main__":

    if len(argv) < 2:
        print("arg1: categories.xlsl")
        exit()

    xlsx_file = argv[1]
    wb=load_workbook(xlsx_file)
    sheet=wb.active

    categories_dict = {}
    ids_dict = {}
    categr_ids = [str(row[0].value) for row in sheet.iter_rows() if row[0].value is not None]
    categr_ids.pop(0)
    categr_ids = [int(val) for val in categr_ids]
    
    parent_ids = [str(row[1].value) for row in sheet.iter_rows() if row[0].value is not None]
    parent_ids.pop(0)
    parent_ids = [int(val) for val in parent_ids]

    categ_names = [str(row[3].value) for row in sheet.iter_rows() if row[0].value is not None]
    categ_names.pop(0)

    for i in range(len(categ_names)):
        if parent_ids[i] == 0:
            categories_dict[categ_names[i]] = categr_ids[i]
            ids_dict[categr_ids[i]] = categ_names[i]
        else:
            current = categ_names[i]
            parent = ids_dict[parent_ids[i]]
            categories_dict[parent + ">" + current] = categr_ids[i]
            ids_dict[categr_ids[i]] = parent + ">" + current

    pprint(categories_dict)

    with open("../pkl_files/categories.pkl", 'wb') as f:
        pickle.dump(categories_dict, f)
