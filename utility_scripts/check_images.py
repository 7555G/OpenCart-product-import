#!/usr/bin/env python

from os import listdir, walk
from os.path import isfile, join
from sys import argv

def replace_chars(chars_string, the_string):
    for char in chars_string:
        the_string = the_string.replace(char, "")

    return the_string

def onlyfiles(mypath):
    return [f for f in listdir(mypath) if isfile(join(mypath, f))]

def clean(path):
    if path[0:2] == "._":
        return path[2:]
    return path

def check_image_exists(product_id, search_path):
    # Recursively check files in folders until you find the image
    dirs = [d for d in listdir(search_path) if not isfile(join(search_path, d))]
    files = onlyfiles(search_path)

    # Attempt No.1: Substrings
    for image in files:
        # Remove spaces
        target_image = image.replace(" ", "")
        target_image = replace_chars("-[+.^:,]_", target_image)

        target_id = product_id.replace(" ", "")
        target_id = replace_chars("-[+.^:,]_", target_id)

        if target_image.lower().find(target_id.lower()) != -1:
            return join(search_path, clean(image))

    # Product not in files of current dir. Check other folders in current dir
    for folder in dirs:
        assoc_image=check_image_exists(product_id, join(search_path, folder))
        if assoc_image:
            return assoc_image

    return 0

def get_ids(xlsx_file, column=0):
    from openpyxl import Workbook, load_workbook

    wb=load_workbook(xlsx_file)
    sheet=wb.active

    return [str(row[column].value) for row in sheet.iter_rows() if row[column].value is not None]
    

if __name__ == "__main__":

    if len(argv) < 3:
        print("Need two args. 1: xlsx file 2: photos dir")
        exit()

    xlsx_file = argv[1]
    search_path = argv[2]

    dic={}
    missing=0
    for product_id in get_ids(xlsx_file):
        associated_image = check_image_exists(product_id, search_path)
        if associated_image:
            dic[product_id]=associated_image
            print(str(product_id) + ": " + associated_image)
        else:
            print(str(product_id) + ": No image found!")
            missing+=1

    print("We are missing: ", missing)
