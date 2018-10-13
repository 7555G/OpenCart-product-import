#!/usr/bin/env python
# -*- coding: utf-8 -*-

from openpyxl import Workbook, load_workbook
from pprint import pprint
from sys import argv
import pickle

if __name__ == "__main__":

    if len(argv) < 2:
        print("arg1: attributes.xlsl")
        exit()

    # load xlsx file
    xlsx_file = argv[1]
    wb=load_workbook(xlsx_file, data_only=True)
    ws=wb["stats"]

    # prepare column name dict
    ncol = ws.max_column
    nrow = ws.max_row
    col = {}
    for i in range(1,ncol + 1):
        col[ws.cell(row = 1, column = i).value.lower()] = i

    data_to_type        = {}
    data_to_subcategory = {}
    for i in range(2, nrow + 1):
        data_to_type[ws.cell(row = i, column = col["type"]).value] = \
            [ws.cell(row = i, column = col["english"]).value, \
             ws.cell(row = i, column = col["greek"]).value]
        data_to_subcategory[ws.cell(row = i, column = col["type"]).value] =\
            ws.cell(row = i, column = col["subcategory"]).value

        with open("../pkl_files/types.pkl", 'wb') as f:
            pickle.dump((data_to_type, data_to_subcategory), f)
