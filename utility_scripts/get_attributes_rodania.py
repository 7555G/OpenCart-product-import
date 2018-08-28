#!/usr/bin/env python

from sys import argv
from pprint import pprint
from openpyxl import Workbook, load_workbook
import csv

PRODUCTS = 67
# Data moving functions
def open_attrs_file(file_path):
    attr_list = open(file_path)
    attr_list = [line.replace('\n','') for line in attr_list]

    # Remove commas and trim spaces
    attr_list = [line.split(',') for line in attr_list]

    for product in range(len(attr_list)):
        for entry in range(len(attr_list[product])):
            attr_list[product][entry] = attr_list[product][entry].strip() 

    # Create attrs dict
    attr_dict = {}
    for i in range(len(attr_list[0])):
        attr_dict[attr_list[0][i]] = attr_list[1][i]

    return attr_dict

def write_data_file(file_path, data):
    file = open(file_path)

def is_column(string):
    if string[0] != '-':
        return 1
    else:
        return 0

def get_data(stock_sheet, product_data, row, attr_dict):
    for item in attr_dict:
        if is_column(attr_dict[item]):
            product_data[row][item] = \
                         stock_sheet[attr_dict[item] + str(row+2)].value
        else:
            product_data[row][item] = attr_dict[item][1:]


if __name__ == '__main__':

    if len(argv) != 4:
        print('arg1: stock.xlsx, arg2: data.in, arg3: attrs.txt')
        exit(1)

    stock_xlsx = argv[1]
    data_in    = argv[2]
    attrs_file = argv[3]

    attr_dict = open_attrs_file(attrs_file)
    wb = load_workbook(stock_xlsx, data_only=True)
    stock_sheet = wb.active
    rows = stock_sheet.max_row
    product_data =[{} for y in range(PRODUCTS)] 

    # Collect the data from the spreadsheet to a variable
    for row in range(PRODUCTS):
        get_data(stock_sheet, product_data, row, attr_dict)
    
    product_data_list = [product.values() for product in product_data] 
    product_data_list.insert(0, product_data[0].keys())

    # Write data to file
    myfile = open(data_in, 'w')
    writer = csv.writer(myfile, dialect='excel', lineterminator='\n')
    writer.writerows(product_data_list) 
