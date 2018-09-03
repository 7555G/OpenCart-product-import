#!/usr/bin/env python

from openpyxl import Workbook, load_workbook
from pprint import pprint
from sys import argv
from distance import levenshtein

def load_categories(file):
    import pickle

    with open(file, 'rb') as f:
        return pickle.load(f)

def load_attributes(file):
    import pickle

    with open(file, 'rb') as f:
        return pickle.load(f)

def replace_chars(chars_string, the_string):
    for char in chars_string:
        the_string = the_string.replace(char, "")

    return the_string

def closest_match(string, strings_container):
    
    smallest_dist = 100
    closest_match = ""
    for match_candidate in strings_container:
    
        new_dist = levenshtein(replace_chars(" >/-._", string).lower(), \
                               replace_chars(" >/-._", match_candidate).lower())    
        if new_dist < smallest_dist:
            smallest_dist = new_dist
            closest_match = match_candidate

    return closest_match

def find_attr_from_category(category):
    (__, attributes_names, attributes_dict) = load_attributes('../pkl_files/attributes.pkl')

    # Convert the greek attribute groups to english
    attributes = [attributes_names[attr] for attr in attributes_dict]    
    attr_grp = closest_match(category, attributes)
    return attributes_names[attr_grp]

if __name__ == "__main__":

    if len(argv) < 2:
        print("arg1: attributes.xlsl")
        exit()

    xlsx_file = argv[1]
    wb=load_workbook(xlsx_file)

    # Attribute Groups
    sheet=wb['AttributeGroups']

    attr_grp_ids = [str(row[0].value) for row in sheet.iter_rows() if row[0].value is not None]
    attr_grp_ids.pop(0)
    attr_grp_ids = [int(row) for row in attr_grp_ids]

    attr_grp_names_el = [str(row[2].value) for row in sheet.iter_rows() if row[0].value is not None]
    attr_grp_names_el.pop(0)
    attr_grp_names_en = [str(row[3].value) for row in sheet.iter_rows() if row[0].value is not None]
    attr_grp_names_en.pop(0)

    # Attributes
    sheet=wb['Attributes']

    attr_ids = [str(row[0].value) for row in sheet.iter_rows() if row[0].value is not None]
    attr_ids.pop(0)
    attr_ids = [int(row) for row in attr_ids]

    parent_ids = [str(row[1].value) for row in sheet.iter_rows() if row[0].value is not None]
    parent_ids.pop(0)
    parent_ids = [int(row) for row in parent_ids]

    attr_names_el = [str(row[3].value) for row in sheet.iter_rows() if row[0].value is not None]
    attr_names_el.pop(0)
    attr_names_en = [str(row[4].value) for row in sheet.iter_rows() if row[0].value is not None]
    attr_names_en.pop(0)

    # For all attr names, match them by language (ΦΥΛΟ:GENDER)
    attr_names = {}
    for i in range(len(attr_grp_names_el)):
        attr_names[attr_grp_names_en[i]] = attr_grp_names_el[i]
        attr_names[attr_grp_names_el[i]] = attr_grp_names_en[i]

    for i in range(len(attr_names_en)):
        attr_names[attr_names_en[i]] = attr_names_el[i]
        attr_names[attr_names_el[i]] = attr_names_en[i]

    # Create the dicts that holds attributes and ids
    attributes_dict = {}    # {'ΡΟΛΟΓΙΑ':['ΦΥΛΟ', 'ΚΑΣΑ'...]}

    # Create the ids dict
    attr_grp_ids_dict = {}    # {'gender':5, ...}
    ids_attr_grp_dict = {}           # {'5':gender, ...}
    for i in range(len(attr_grp_ids)):
        ids_attr_grp_dict[attr_grp_ids[i]] = attr_grp_names_el[i]
        attr_grp_ids_dict[attr_grp_names_el[i]] = attr_grp_ids[i]

    attr_ids_dict = {}    # {'gender':5, ...}
    ids_attr_dict = {}           # {'5':gender, ...}
    for i in range(len(attr_ids)):
        ids_attr_dict[attr_ids[i]] = attr_names_el[i]
        attr_ids_dict[attr_names_el[i]] = attr_ids[i]

    # Initialize the attributes_dict
    for attr_grp in attr_grp_names_el:
        attributes_dict[attr_grp] = []

    # Create the list of attributes for each attribute group
    for i in range(len(parent_ids)):
        attr = attr_names_el[i]
        attr_grp = ids_attr_grp_dict[parent_ids[i]]
        attributes_dict[attr_grp].append(attr)


    # pprint(attributes_dict)

    categories_to_attributes = {}
    with open("../pkl_files/attributes.pkl", 'wb') as f:
        pickle.dump((categories_to_attributes, attr_names, attributes_dict), f)
        
    # We will hand craft a dictionary that has the matchup between categories and attr groups
    categories_dict = load_categories('../pkl_files/categories.pkl')
    # print("categories_to_attributes = {")
    # for categ in categories_dict:
    #     print("\t'" + categ + "': '" + find_attr_from_category(categ) + "',")
    # print("\t}")
    
    for categ in categories_dict:
        categories_to_attributes[categ]= find_attr_from_category(categ)

    # Fix by hand errors
    categories_to_attributes['WATCHES>SMARTWATCHES'] = 'ΡΟΛΟΓΙΑ'

    # Based from the output of the commented commands we have something like this
    # categories_to_attributes = {}
    pprint(attributes_dict)

    with open("../pkl_files/attributes.pkl", 'wb') as f:
        pickle.dump((categories_to_attributes, attr_names, attributes_dict), f)
