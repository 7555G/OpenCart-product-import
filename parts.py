#!/usr/bin/env python

from sys import argv
from pprint import pprint
from openpyxl import Workbook, load_workbook

from libs.transformations import COLUMN_TRANSF_RULES
from libs.color import get_color
from libs.utilities import *

MANUFACTURER = "Maurice Lacroix"
CATEG = "WATCHES>WATCH SPARE PARTS"
OFFER_CATEG = "OFFERS>WATCH SPARE PARTS"


def cleanup(wb):
    products_sheet = wb['Products']
    products_sheet.append(['' for i in range(products_sheet.max_column)])
    row_num = products_sheet.max_row

    # print("Row to clean is: ", row_num)
    products_sheet.delete_rows(row_num)

def process_attr_data(product, attr_grp):
    for attr in COLUMN_TRANSF_RULES[attr_grp]:
        # First create the list(of lists) based on the priorities
        transformations_ordered = [[] for i in range(len(COLUMN_TRANSF_RULES[attr_grp][attr]) + 1000)] 
        
        for word, transf in COLUMN_TRANSF_RULES[attr_grp][attr].items():
            transformations_ordered[transf[-1]].append(word)
            # [ ['STEEL','S.S'], ['Ceramic'], ... ]
            
        # Clean empty lists
        transformations_ordered = [lst for lst in transformations_ordered if lst]
        # Watch out for no data
        try:
            if not product[attr]:
                product[attr] = ""
                continue
        except:
            product[attr] = ""


        new_val_el = ""
        new_val_en = ""
        for transf in transformations_ordered:
            possible_words = [word for word in transf]

            max_word = ""
            # Now check if any of the words are in the text
            for word in possible_words:

                if "__color" in word:
                    
                    colors_vals = get_color(word.lower(), replace_chars("/|", product[attr].lower(), " "))
                    if colors_vals:
                        if len(word) > len(max_word):
                            max_word = word

                elif word.upper() in product[attr].upper():
                    if len(word) > len(max_word):
                        max_word = word

            # If a word was matched, it will be the one with the most characters
            if max_word:
                new_val_en += ' ' + COLUMN_TRANSF_RULES[attr_grp][attr][max_word][0]
                new_val_el += ' ' + COLUMN_TRANSF_RULES[attr_grp][attr][max_word][1]

                # If the pattern matched we have the colors that need to be replaced
                if "__color" in max_word:
                    colors_vals = get_color(max_word.lower(), replace_chars("/|", product[attr].lower(), " "))
                    new_val_el = new_val_el.replace("__color", colors_vals[1])
                    new_val_en = new_val_en.replace("__color", colors_vals[0])
        
        if not new_val_el:
            product[attr] = [product[attr], product[addr]]
        elif new_val_el == " __clear":
            product[attr] = ["", ""]
        else:
            new_val_el = new_val_el[1:][0].upper() + new_val_el[2:]
            new_val_en = new_val_en[1:][0].upper() + new_val_en[2:]
            if new_val_el[-1] == ',':
                new_val_el = new_val_el[:-1]
                new_val_en = new_val_en[:-1]
            product[attr] = [new_val_el, new_val_en]

    # Lastly all attributes that weren't filtered, convert them to [greek, engl] format
    for attr in product:
        if not isinstance(product[attr], list) and attr in COLUMN_TRANSF_RULES[attr_grp]:
            product[attr] = [product[attr], product[attr]]

    return product        

def static_pre_processing(product_info, attr_grp):
    # Apply rules to the data before inserting it
    # product_info['name'] = product_info['name'].title()
    product_info['ΥΛΙΚΟ'] = product_info['name']
    product_info['ΧΡΩΜΑ'] = product_info['name']
    return product_info

def static_post_processing(product_info, attr_grp):
    (types, __) = load_pickle_obj("pkl_files/types.pkl")
    
    product_info['ΕΙΔΟΣ'] = ["",""]
    product_info['ΕΙΔΟΣ'][0] = types[product_info['type']][1].title()
    product_info['ΕΙΔΟΣ'][1] = types[product_info['type']][0].title()
    product_info['type'] = types[product_info['type']]
    product_info['ΠΛΗΡΟΦΟΡΙΕΣ'] = ["", ""]

    if product_info["used"]:
        product_info['ΠΛΗΡΟΦΟΡΙΕΣ'][0] = "Χρησιμοποιείται στα "+product_info["used"]
        product_info['ΠΛΗΡΟΦΟΡΙΕΣ'][1] = "Used in "+product_info["used"]

    return product_info

def open_new_products(input_file):
    new_products = open(input_file)
    new_products = [line.replace('\n','') for line in new_products]

    # Remove commas and trim spaces
    new_products = [line.split(',') for line in new_products]

    for product in range(len(new_products)):
        for entry in range(len(new_products[product])):
            new_products[product][entry] = \
                                        new_products[product][entry].strip()

    # Format data
    for product in range(len(new_products)):
        # Manufacturer
        new_products[product][COL["number"]] = \
                                   new_products[product][COL["number"]].title()
        # Category
        new_products[product][CATEG_INDX] = \
                          new_products[product][CATEG_INDX].replace(' ', '')
        new_products[product][CATEG_INDX] = \
                                   new_products[product][CATEG_INDX].upper()
        # Gender
        new_products[product][GENDER_INDX] = \
                                  new_products[product][GENDER_INDX].lower()
        # Price
        new_products[product][PRICE_INDX] = \
                                           new_products[product][PRICE_INDX]

        # Family
        new_products[product][FAMILY_INDX] = \
                                  new_products[product][FAMILY_INDX].upper()

        # Collection
        new_products[product][COLLEC_INDX] = \
                                  new_products[product][COLLEC_INDX].title()

    return new_products

def open_product_attributes(input_file):
    # The tsv will be converted to a list of dictionaries
    # Now we index as [prod_num]["type"]
    new_attrs = open(input_file)
    new_attrs = [line.replace('\n','') for line in new_attrs] 
    new_attrs = [line.split('\t') for line in new_attrs]
    attr_names = new_attrs.pop(0)

    attrs_dicts = [{} for prod in new_attrs]
    for prod_ind in range(len(new_attrs)):
        for i in range(len(new_attrs[prod_ind])):
            attrs_dicts[prod_ind][attr_names[i]] = new_attrs[prod_ind][i]
    
    return attrs_dicts

def load_pickle_obj(file):
    import pickle

    with open(file, 'rb') as f:
        return pickle.load(f)

# XLSX modifier functions
def add_empty_product(product_info, wb):
    products_sheet = wb['Products']
    products_sheet.append(['' for i in range(products_sheet.max_column)])
    row_num = products_sheet.max_row

    # Write the new product code
    last_product_id = products_sheet['A' + str(row_num - 1)].value
    curr_product_id = last_product_id + 1
    products_sheet['A' + str(row_num)] = curr_product_id
    
    print("Insert new product with ID {} in row {} with Model: {}".format( \
        curr_product_id, row_num, product_info["number"]))

def add_attributes(product_info, wb):
    products_sheet = wb['Products']
    attributes_sheet = wb['ProductAttributes']
    row_num = products_sheet.max_row
    attr_row_num = attributes_sheet.max_row + 1

    (categories_to_attribute, __, attributes_dict) = load_pickle_obj('pkl_files/attributes.pkl')
    
    attr_grp = categories_to_attribute[CATEG]
    attributes = attributes_dict[attr_grp]

    last_product_id = products_sheet['A' + str(row_num - 1)].value
    curr_product_id = last_product_id + 1

    # Data Processing
    attribute_info = static_pre_processing(product_info, attr_grp)
    attribute_info = process_attr_data(attribute_info, attr_grp)
    attribute_info = static_post_processing(product_info, attr_grp)
    
    i = 0
    for attr in attributes:
        # If attribute is empty, don't insert the row
        if attr not in attribute_info or attribute_info[attr][0] == "":
            continue
        attributes_sheet.append(['' for j in range(attributes_sheet.max_column)])
        
        attributes_sheet['C' + str(attr_row_num + i)] = attr
        attributes_sheet['A' + str(attr_row_num + i)] = curr_product_id
        attributes_sheet['B' + str(attr_row_num + i)] = attr_grp

        if attr in attribute_info:
            attributes_sheet['D' + str(attr_row_num + i)] = attribute_info[attr][0]
            attributes_sheet['E' + str(attr_row_num + i)] = attribute_info[attr][1]
        else:
            attributes_sheet['D' + str(attr_row_num + i)] = ''
            attributes_sheet['E' + str(attr_row_num + i)] = ''
        i += 1

def add_product_name(product_info, wb):
    products_sheet = wb['Products']
    row_num = products_sheet.max_row

    # Create product name
    product_name = product_info["name"]

    # Write product name
    products_sheet['B' + str(row_num)] = product_name
    products_sheet['C' + str(row_num)] = product_name

def add_description(product_info, wb):
    products_sheet = wb['Products']
    row_num = products_sheet.max_row
    descr_el = descr_en = descr_meta_el = descr_meta_en = product_info['name']

    if product_info["used"]:
        descr_el += ".</p><p>\n" + "Χρησιμοποιείται στα "
        descr_en += ".</p><p>\n" + "Used in "
        descr_meta_el += ". Χρησιμοποιείται στα "
        descr_meta_en += ". Used in "
    descr_el += product_info["used"] 
    descr_en += product_info["used"]
    descr_meta_el += product_info["used"]
    descr_meta_en += product_info["used"]

    # Write description
    products_sheet['AE' + str(row_num)] = '<p>' + descr_el + '</p>'
    products_sheet['AF' + str(row_num)] = '<p>' + descr_en + '</p>'

    # Write meta description
    products_sheet['AI' + str(row_num)] = descr_meta_el
    products_sheet['AJ' + str(row_num)] = descr_meta_en

def add_SEO(product_info, wb):
    products_sheet = wb['Products']
    row_num = products_sheet.max_row

    SEO = "Maurice_Lacroix-" + product_info["number"]
    # Write SEO
    products_sheet['AD' + str(row_num)] = SEO

def add_model(product_info, wb):
    products_sheet = wb['Products']
    row_num = products_sheet.max_row

    # Get model
    model = product_info["number"]

    # Write model
    products_sheet['M' + str(row_num)] = model

def add_meta_title(product_info, wb):
    products_sheet = wb['Products']
    row_num = products_sheet.max_row

    meta_title_en  = "Maurice Lacroix " + product_info["ΕΙΔΟΣ"][1] 
    meta_title_en += " " + product_info["number"]

    meta_title_el  = "Maurice Lacroix " + product_info["ΕΙΔΟΣ"][0] 
    meta_title_el += " " + product_info["number"]

    # Wrte meta titles
    products_sheet['AG' + str(row_num)] = meta_title_el
    products_sheet['AH' + str(row_num)] = meta_title_en

def add_price(product_info, wb):
    products_sheet = wb['Products']
    row_num = products_sheet.max_row

    # Get price
    price = product_info["price"]

    # Write price
    products_sheet['Q' + str(row_num)] = price

def add_manufacturer(product_info, wb):
    products_sheet = wb['Products']
    row_num = products_sheet.max_row

    # Check manufacturer
    manuf = 'Maurice Lacroix'
    if manuf == '':
        print('Warning: invalid manufacturer name!')

    # Write manufacturer
    products_sheet['N' + str(row_num)] = manuf

def add_category(product_info, wb):
    products_sheet = wb['Products']
    row_num = products_sheet.max_row
    categories_dict = load_pickle_obj('pkl_files/categories.pkl')

    # Find category number from dictionary
    categ = closest_match(CATEG, categories_dict)

    # Also add the parent categories
    broken_category = categ.split(">")
    parent_categ = broken_category[0]
    categ_val = str(categories_dict[parent_categ])
    for i in range(1, len(broken_category)):
        parent_categ = parent_categ + '>' + broken_category[i]
        categ_val += "," + str( categories_dict[parent_categ] )

    # Write category number
    discount = int('0' + str(product_info['discount']).replace('-', ''))
    if discount != 0:
        categ_val += "," + str( categories_dict["OFFERS"])
        categ_val += "," + str( categories_dict[OFFER_CATEG])

    products_sheet['D' + str(row_num)] = categ_val

def add_status(product_info, wb):
    products_sheet = wb['Products']
    row_num = products_sheet.max_row

    if not product_info['hidden']:
        products_sheet['AB' + str(row_num)] = 'true'
    else:
        products_sheet['AB' + str(row_num)] = 'false'

def add_image(product_info, wb):
    products_sheet = wb['Products']
    row_num = products_sheet.max_row
    alph = 'abcdefghijklmnopqrstuvwxyz'
    image_dir = 'catalog/product/' + CATEG.replace('>', '/') \
              + '/' + product_info["number"] + ".jpg"

    # Write image directory
    if int('0' + str(product_info['img count'])) == 0:
        # print(product_info['img count'] + ' is 0')
        products_sheet['O' + str(row_num)] = 'catalog/product/placeholder.jpg'
        return

    if int(product_info['img count']) == 1:
        products_sheet['O' + str(row_num)] = image_dir
        return

    products_sheet['O' + str(row_num)] = image_dir[:-4] + 'a.jpg'

    # Extra images
    sheet = wb["AdditionalImages"]

    for i in range(1, int(product_info['img count'])):
        sheet.append(['' for i in range(sheet.max_column)])
        row = sheet.max_row
        last_product_id = products_sheet['A' + str(row_num - 1)].value
        curr_product_id = last_product_id + 1
        image_dir = 'catalog/product/' + CATEG.replace('>', '/') \
                  + '/' + product_info["number"] + alph[i] + ".jpg"
        sheet['A' + str(row)] = curr_product_id
        sheet['B' + str(row)] = image_dir
        sheet['C' + str(row)] = i

def add_discount(product_info, wb):
    discount = int('0' + str(product_info['discount']).replace('-', ''))
    if discount == 0: return
    price    = int(product_info['price'])

    products_sheet = wb['Products']
    discounts_sheet = wb['Specials']
    row = discounts_sheet.max_row + 1
    row_num = products_sheet.max_row
    last_product_id = products_sheet['A' + str(row_num - 1)].value
    curr_product_id = last_product_id + 1
    discounts_sheet.append(['' for i in range(discounts_sheet.max_column)])

    discounts_sheet['A' + str(row)] = curr_product_id
    discounts_sheet['B' + str(row)] = 'Default'
    discounts_sheet['C' + str(row)] = '0'
    discounts_sheet['D' + str(row)] = price*(1 - discount/100)
    discounts_sheet['E' + str(row)] = '0000-00-00'
    discounts_sheet['F' + str(row)] = '0000-00-00'

def add_misc(product_info, wb):
    products_sheet = wb['Products']
    row_num = products_sheet.max_row

    
    products_sheet['L'  + str(row_num)] = product_info["stock"]      #quantity
    products_sheet['P'  + str(row_num)] = 'yes'  #shipping
    products_sheet['R'  + str(row_num)] = 0      #points
    products_sheet['S'  + str(row_num)] = '2018-09-11 14:00:00' #date_added
    products_sheet['T'  + str(row_num)] = '2018-09-11 14:00:00'#date_modified
    products_sheet['U'  + str(row_num)] = '2018-09-11' #date_available
    products_sheet['V'  + str(row_num)] = 0      #weight
    products_sheet['W'  + str(row_num)] = 'kg'   #weight_unit
    products_sheet['X'  + str(row_num)] = 0      #length
    products_sheet['Y'  + str(row_num)] = 0      #width
    products_sheet['Z'  + str(row_num)] = 0      #height
    products_sheet['AA' + str(row_num)] = 'cm'   #length_unit
    products_sheet['AC' + str(row_num)] = 0      #tax_class_id
    products_sheet['AM' + str(row_num)] = 6      #stock_status_id
    products_sheet['AN' + str(row_num)] = 0      #store_ids
    products_sheet['AS' + str(row_num)] = 1      #sort_order
    products_sheet['AT' + str(row_num)] = 'true' #subtract
    products_sheet['AU' + str(row_num)] = 1      #minimum


if __name__ == '__main__':

    if len(argv) < 3:
        print('arg1: specs.tsv | arg2: products.xlsx')
        exit(1)

    SPECS_TSV = argv[1]
    products_xlsx = argv[2]

    wb = load_workbook(products_xlsx)
    new_products = open_product_attributes(SPECS_TSV)
    products = 0
    
    # Iterate the inputs
    for product in new_products:
        if not product["type"]: continue
        add_empty_product(product, wb)
        add_attributes(product, wb)
        add_product_name(product, wb)
        add_description(product, wb)
        add_SEO(product, wb)
        add_model(product, wb)
        add_meta_title(product, wb)
        add_price(product, wb)
        add_manufacturer(product, wb)
        add_category(product, wb)
        add_status(product, wb)
        add_image(product, wb)
        add_discount(product, wb)
        add_misc(product, wb)
        products += 1

    # Cleanup and Save to file
    print("Added {} products.".format(products))
    cleanup(wb)
    wb.save(products_xlsx)
