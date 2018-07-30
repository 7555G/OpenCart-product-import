#!/usr/bin/env python

from sys import argv
from pprint import pprint
from openpyxl import Workbook, load_workbook
from distance import levenshtein

# Global Data
<<<<<<< HEAD
=======
categories_dict = {}

>>>>>>> 76c86cbe50c21b0d9235012a65ddbc08f1a5876d
MODEL_INDX  = 0
MANUF_INDX  = 1
CATEG_INDX  = 2
COLLEC_INDX = 3
FAMILY_INDX = 4
GENDER_INDX = 5
PRICE_INDX  = 6

MANUFACTURER = ['Chronoswiss',
                'Fortis',
                'Jos Von Arx',
                'Jowissa',
                'Manfred Cracco',
                'Rodania',
                'Victorinox']

DEFAULT_ATTR = [['n', 'τ'],
                ['o', 'ι'],
                ['t', 'ι'],
                ['h', 'ι'],
                ['i', 'ι'],
                ['n', 'ι'],
                ['g', 'ι'],
                [' ', 'ι'],
                ['t', 'ι'],
                ['o', 'ι'],
                [' ', 'ι'],
                ['s', 'ι'],
                ['e', 'ι'],
                ['e', 'ι'],
                ['e', 'ι']]


def open_new_products(input_file):
    new_products = open(input_file)
    new_products = [line.replace('\n','') for line in new_products]

    # Remove commas and trim spaces
    new_products = [line.split(',') for line in new_products]

    for product in range(len(new_products)):
        for entry in range(len(new_products[product])):
            new_products[product][entry] = \
                                        new_products[product][entry].strip()

    # Format data
    for product in range(len(new_products)):
        # Manufacturer
        new_products[product][MANUF_INDX] = \
                                   new_products[product][MANUF_INDX].title()
        # Category
        new_products[product][CATEG_INDX] = \
                          new_products[product][CATEG_INDX].replace(' ', '')
        new_products[product][CATEG_INDX] = \
                                   new_products[product][CATEG_INDX].upper()
        # Gender
        new_products[product][GENDER_INDX] = \
                                  new_products[product][GENDER_INDX].lower()
        # Price
        new_products[product][PRICE_INDX] = \
                                      int(new_products[product][PRICE_INDX])
    
    return new_products

def load_pickle_obj(file):
    import pickle

    with open(file, 'rb') as f:
        return pickle.load(f)

def replace_chars(chars_string, the_string):
    for char in chars_string:
        the_string = the_string.replace(char, "")

    return the_string

def closest_match(string, strings_container):
    
    smallest_dist = 100
    closest_match = ""
    for match_candidate in strings_container:
    
        new_dist = levenshtein(replace_chars(" >/-._", string).lower(), \
                               replace_chars(" >/-._", match_candidate).lower())    
        if new_dist < smallest_dist:
            smallest_dist = new_dist
            closest_match = match_candidate

    return closest_match


# XLSX modifier functions
def add_empty_product(product_info, wb):
    products_sheet = wb['Products']
    attributes_sheet = wb['ProductAttributes']
    products_sheet.append(['' for i in range(products_sheet.max_column)])
    row_num = products_sheet.max_row
    attr_row_num = attributes_sheet.max_row + 1

    (categories_to_attribute, attributes_dict) = load_pickle_obj('pkl_files/attributes.pkl')
    
    attr_grp = categories_to_attribute[product_info[CATEG_INDX]]
    attributes = attributes_dict[attr_grp]
    
    for i in range(len(attributes)):
        attributes_sheet.append(['' \
                              for j in range(attributes_sheet.max_column)])
   
    # Write the new product code
    last_product_id = products_sheet['A' + str(row_num - 1)].value
    curr_product_id = last_product_id + 1
    products_sheet['A' + str(row_num)] = curr_product_id
    for i in range(len(attributes)):
        attributes_sheet['C' + str(attr_row_num + i)] = attributes[i]
        attributes_sheet['A' + str(attr_row_num + i)] = curr_product_id
        attributes_sheet['B' + str(attr_row_num + i)] = attr_grp
        attributes_sheet['D' + str(attr_row_num + i)] = DEFAULT_ATTR[i][0]
        attributes_sheet['E' + str(attr_row_num + i)] = DEFAULT_ATTR[i][1]

    print('Insert new product with ID ' + str(curr_product_id) \
          + ' in row ' + str(row_num) + '.')


def add_product_name(product_info, wb):
    products_sheet = wb['Products']
    row_num = products_sheet.max_row

    # Create product name
    product_name = product_info[MANUF_INDX]  + ' ' \
                 + product_info[COLLEC_INDX] + ' '
    if product_info[FAMILY_INDX] != '':
        product_name += product_info[FAMILY_INDX] + ' '
    product_name += product_info[MODEL_INDX]

    # Write product name
    products_sheet['B' + str(row_num)] = product_name
    products_sheet['C' + str(row_num)] = product_name


def add_description(product_info, wb):
    products_sheet = wb['Products']
    row_num = products_sheet.max_row

    # Define gender
    if product_info[GENDER_INDX] == 'mens':
        gender_en = 'mens'
        gender_el = 'ανδρικό'
    elif product_info[GENDER_INDX] == 'womens':
        gender_en = 'womens'
        gender_el = 'γυναικείο'

    # Greek
    descr_el = 'Κομψό ' + gender_el + ' ρολόι από την εταιρία ' \
             + product_info[MANUF_INDX] \
             + ', Ελβετικής κασκευής, από τη συλλογή ' \
             + product_info[COLLEC_INDX] 
    if product_info[FAMILY_INDX] != '':
        # Family could be empty
        descr_el += ' ' + product_info[FAMILY_INDX]
    descr_el += ', με κωδικό ' + product_info[MODEL_INDX] + '.'

    # English
    descr_en = 'An elegant ' + gender_en + ' watch by ' \
             + product_info[MANUF_INDX] \
             + ', Swiss made, from the collection ' \
             + product_info[COLLEC_INDX] 
    if product_info[FAMILY_INDX] != '':
        # Family could be empty
        descr_en += ' ' + product_info[FAMILY_INDX]
    descr_en += ', with reference ' + product_info[MODEL_INDX] + '.'

    # Write description
    products_sheet['AE' + str(row_num)] = '<p>' + descr_el + '</p>'
    products_sheet['AF' + str(row_num)] = '<p>' + descr_en + '</p>'

    # Write meta description
    products_sheet['AI' + str(row_num)] = descr_el
    products_sheet['AJ' + str(row_num)] = descr_en


def add_SEO(product_info, wb):
    products_sheet = wb['Products']
    row_num = products_sheet.max_row

    # Create SEO
    SEO = product_info[MANUF_INDX].replace(' ', '_') + '-' \
        + product_info[COLLEC_INDX].replace(' ', '_') + '-'
    if product_info[FAMILY_INDX] != '':
        SEO += product_info[FAMILY_INDX].replace(' ', '_') + '-'
    SEO += product_info[MODEL_INDX].replace(' ', '_')

    # Write SEO
    products_sheet['AD' + str(row_num)] = SEO


def add_model(product_info, wb):
    products_sheet = wb['Products']
    row_num = products_sheet.max_row

    # Get model
    model = product_info[MODEL_INDX]

    # Write model
    products_sheet['M' + str(row_num)] = model


def add_meta_title(product_info, wb):
    products_sheet = wb['Products']
    row_num = products_sheet.max_row

    # Define gender
    if product_info[GENDER_INDX] == 'mens':
        gender_en = 'Mens'
        gender_el = 'Ανδρικό'
    elif product_info[GENDER_INDX] == 'womens':
        gender_en = 'Womens'
        gender_el = 'Γυναικείο'

    # Greek
    if product_info[FAMILY_INDX] != '':
        meta_title_el_m = gender_el + ' Ελβετικό ρολόι - ' \
                        + product_info[MANUF_INDX]  + ' ' \
                        + product_info[FAMILY_INDX] + ' ' \
                        + product_info[MODEL_INDX]  + ' | Eurotimer'
        meta_title_el_l = gender_el + ' Ελβετικό ρολόι - ' \
                        + product_info[MANUF_INDX]  + ' ' \
                        + product_info[COLLEC_INDX] + ' ' \
                        + product_info[FAMILY_INDX] + ' ' \
                        + product_info[MODEL_INDX]  + ' | Eurotimer'
    else:
        meta_title_el_m = gender_el + ' Ελβετικό ρολόι - ' \
                        + product_info[MANUF_INDX]  + ' ' \
                        + product_info[COLLEC_INDX] + ' ' \
                        + product_info[MODEL_INDX]  + ' | Eurotimer'
        meta_title_el_l = meta_title_el_m
    meta_title_el_s = gender_el + ' Ελβετικό ρολόι - ' \
                    + product_info[MANUF_INDX] + ' ' \
                    + product_info[MODEL_INDX] + ' | Eurotimer'
    
    # English
    if product_info[FAMILY_INDX] != '':
        meta_title_en_m = gender_en + ' Watches - Swiss Made ' \
                        + product_info[MANUF_INDX]  + ' ' \
                        + product_info[FAMILY_INDX] + ' ' \
                        + product_info[MODEL_INDX]  + ' | Eurotimer'
        meta_title_en_l = gender_en + ' Watches - Swiss Made ' \
                        + product_info[MANUF_INDX]  + ' ' \
                        + product_info[COLLEC_INDX] + ' ' \
                        + product_info[FAMILY_INDX] + ' ' \
                        + product_info[MODEL_INDX]  + ' | Eurotimer'
    else:
        meta_title_en_m = gender_en + ' Watches - Swiss Made ' \
                        + product_info[MANUF_INDX]  + ' ' \
                        + product_info[COLLEC_INDX] + ' ' \
                        + product_info[MODEL_INDX]  + ' | Eurotimer'
        meta_title_en_l = meta_title_en_m
    meta_title_en_s = gender_en + ' Watches - Swiss Made ' \
                    + product_info[MANUF_INDX] + ' ' \
                    + product_info[MODEL_INDX] + ' | Eurotimer'

    # Select meta title with appropriate length
    if len(meta_title_el_l) <= 60:
        meta_title_el = meta_title_el_l
    elif len(meta_title_el_m) <= 60:
        meta_title_el = meta_title_el_m
    elif len(meta_title_el_s) <= 60:
        meta_title_el = meta_title_el_s
    else:
        print('Warning: Greek meta title is longer than 60 characters!')
        meta_title_el = ''

    if len(meta_title_en_l) <= 60:
        meta_title_en = meta_title_en_l
    elif len(meta_title_en_m) <= 60:
        meta_title_en = meta_title_en_m
    elif len(meta_title_en_s) <= 60:
        meta_title_en = meta_title_en_s
    else:
        print('Warning: English meta title is longer than 60 characters!')
        meta_title_en = ''

    # Wrte meta titles
    products_sheet['AG' + str(row_num)] = meta_title_el
    products_sheet['AH' + str(row_num)] = meta_title_en


def add_price(product_info, wb):
    products_sheet = wb['Products']
    row_num = products_sheet.max_row

    # Get price
    price = product_info[PRICE_INDX]

    # Write price
    products_sheet['Q' + str(row_num)] = price


def add_manufacturer(product_info, wb):
    products_sheet = wb['Products']
    row_num = products_sheet.max_row

    # Check manufacturer
    manuf = ''
    for correct_manuf in MANUFACTURER:
        if product_info[MANUF_INDX]== correct_manuf:
            manuf = product_info[MANUF_INDX]
    if manuf == '':
        print('Warning: invalid manufacturer name!')

    # Write manufacturer
    products_sheet['N' + str(row_num)] = manuf


def add_category(product_info, wb):
    products_sheet = wb['Products']
    row_num = products_sheet.max_row
    categories_dict = load_pickle_obj('pkl_files/categories.pkl')

    # Find category number from dictionary
    categ = closest_match(product_info[CATEG_INDX], categories_dict)

    # Write category number
    products_sheet['D' + str(row_num)] = categories_dict[categ]


def add_image(product_info, wb):
    products_sheet = wb['Products']
    row_num = products_sheet.max_row

    # Define image directory
    model = product_info[MODEL_INDX]
    categ_dir = product_info[CATEG_INDX].replace('>', '/')
    image_dir = 'catalog/product/' + categ_dir + '/' + model + '.jpg'

    # Write image directory
    products_sheet['O' + str(row_num)] = image_dir


def add_misc(product_info, wb):
    products_sheet = wb['Products']
    row_num = products_sheet.max_row

    
    products_sheet['L'  + str(row_num)] = 1      #quantity
    products_sheet['P'  + str(row_num)] = 'yes'  #shipping
    products_sheet['R'  + str(row_num)] = 0      #points
    products_sheet['S'  + str(row_num)] = '2018-07-26 14:00:00' #date_added
    products_sheet['T' + str(row_num)] = '2018-07-26 14:00:00'#date_modified
    products_sheet['U'  + str(row_num)] = '2018-07-26' #date_available
    products_sheet['V'  + str(row_num)] = 0      #weight
    products_sheet['W'  + str(row_num)] = 'kg'   #weight_unit
    products_sheet['X'  + str(row_num)] = 0      #length
    products_sheet['Y'  + str(row_num)] = 0      #width
    products_sheet['Z'  + str(row_num)] = 0      #height
    products_sheet['AA' + str(row_num)] = 'cm'   #length_unit
    products_sheet['AB' + str(row_num)] = 'true' #status
    products_sheet['AC' + str(row_num)] = 0      #tax_class_id
    products_sheet['AM' + str(row_num)] = 6      #stock_status_id
    products_sheet['AN' + str(row_num)] = 0      #store_ids
    products_sheet['AS' + str(row_num)] = 1      #sort_order
    products_sheet['AT' + str(row_num)] = 'true' #subtract
    products_sheet['AU' + str(row_num)] = 1      #minimum


if __name__ == '__main__':

    if len(argv) != 3:
        print('arg1: input_file, arg2: products.xlsx')
        exit(1)

    input_file = argv[1]
    products_xlsx = argv[2]

    wb = load_workbook(products_xlsx)
    new_products = open_new_products(input_file)

    # Iterate the inputs
    for product in new_products:
        add_empty_product(product, wb)
        add_product_name(product, wb)
        add_description(product, wb)
        add_SEO(product, wb)
        add_model(product, wb)
        add_meta_title(product, wb)
        add_price(product, wb)
        add_manufacturer(product, wb)
        add_category(product, wb)
        add_image(product, wb)
        add_misc(product, wb)

    # Save to file
    wb.save(products_xlsx)
