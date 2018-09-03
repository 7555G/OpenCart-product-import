#!/usr/bin/env python
from sys import argv
from openpyxl import Workbook, load_workbook
from pprint import pprint
from selenium.webdriver.firefox.options import Options
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains


# setup browser
options = Options()
options.add_argument("--headless")
browser = webdriver.Firefox(firefox_options=options)

def site_login(credentials_txt):
    with open(credentials_txt) as f:
        username = f.readline().rstrip()
        password = f.readline().rstrip()
    browser.get("http://extranet.mauricelacroix.com/")
    browser.find_element_by_id("username") \
           .send_keys(username)
    browser.find_element_by_id ("password") \
           .send_keys(password)
    browser.find_element_by_class_name("darkButton") \
           .click()

def goto_part_search_by_ref(reference):
    ref_search_prefix = \
        "http://extranet.mauricelacroix.com/sav/components/list.html?reference="
    browser.get(ref_search_prefix + reference)

def goto_part_search_by_key(keyword):
    key_search_prefix = \
        "http://extranet.mauricelacroix.com/sav/components/list.html?keywords="
    browser.get(key_search_prefix + keyword)

def goto_part_sheet(sheet):
    part_sheet_prefix = \
        "http://extranet.mauricelacroix.com/sav/components/sheet_"
    part_sheet_suffix = ".html"
    browser.get(part_sheet_prefix + sheet + part_sheet_suffix)

def goto_used_with(sheet):
    used_with_prefix = \
        "http://extranet.mauricelacroix.com/sav/watches/list.html?usecase="
    browser.get(used_with_prefix + sheet)

def fetch_sheet(n):
    # from search page
    n = str(n)
    sheet_xpath = "//div[" + n + "]/div[5]/div/div[1]/input"
    try:
        sheet = browser.find_element_by_xpath(sheet_xpath) \
                       .get_attribute("value")
    except:
        return ""
    sheet = sheet[:-5].split("_")[-1]
    return sheet

def fetch_type():
    # from part sheet page
    type_xpath = "/html/body/div[1]/div[@class='watchfile_title  ']"
    try:
        part_type = browser.find_element_by_xpath(type_xpath) \
                           .text
    except:
        return ""
    return part_type

def fetch_name():
    # from part sheet page
    name_xpath = "/html/body/div[2]/div[1][@class='regular_header darker']"
    try:
        name = browser.find_element_by_xpath(name_xpath) \
                      .text
    except:
        return ""
    return name

def fetch_number():
    # from part sheet page
    number_xpath = "/html/body/div[2]/div[2]/span[2][@class='dark font90']"
    try:
        number = browser.find_element_by_xpath(number_xpath) \
                        .text
    except:
        return ""
    return number

def fetch_info():
    # from part sheet page
    info_xpath = "/html/body/div[2]/div[2]/span[3][@class='darker font95']"
    try:
        info = browser.find_element_by_xpath(info_xpath) \
                      .text
    except:
        return ""
    return info

def download_image(name):
    # from part sheet page
    try:
        img_url = browser.find_element_by_xpath("/html/body/div[1]/img") \
                         .get_attribute("src")
    except:
        return ""
    if "placeholder" in img_url: return ""

    #browser.get(img_url)
    #image_xpath = "/html/body"
    #ActionChains(browser) \
    #       .key_down(Keys.CONTROL) \
    #       .send_keys("s") \
    #       .key_up(Keys.CONTROL) \
    #       .perform()
    return img_url

def fetch_used_with():
    # from "used with" page
    used_with_xpath = "//div[@class='watchlist-container mb20']"
    try:
        watches_obj = browser.find_elements_by_xpath(used_with_xpath)
    except:
        return ""
    if not watches_obj:
        return ""
    used_with = [x.text for x in watches_obj]
    used_with = used_with[0].split("\n")
    used_with = [x.strip(" ") for x in used_with]
    used_with = " ".join(used_with)
    return used_with


if __name__ == "__main__":

    if len(argv) < 3:
        print("arg1: credentials.txt, arg2: parts_stock.xlsl")
        exit()

    # load xlsx file
    xlsx_file = argv[2]
    wb=load_workbook(xlsx_file)
    ws=wb["attributes"]

    # prepare column name dict
    ncol = ws.max_column
    nrow = ws.max_row
    col = {}
    for i in range(1,ncol + 1):
        col[ws.cell(row = 1, column = i).value] = i

    # login to site
    site_login(argv[1])

    for i in range(2, nrow + 1):
        # show progress
        print("item: " + str(i - 2) + "/" + str(nrow + 1), end="\r")

        part_code = ws.cell(row = i, column = col["part code"]).value

        # go to part search and fetch 1st part's sheet number
        goto_part_search_by_ref(part_code)
        sheet = fetch_sheet(1)
        if sheet == "":
            goto_part_search_by_key(part_code)
            sheet = fetch_sheet(1)

        if sheet == "":
            print("no results for " + part_code)
            continue

        # go to part's sheet page and fetch data
        goto_part_sheet(sheet)
        number    = fetch_number()
        part_type = fetch_type()
        name      = fetch_name()
        info      = fetch_info()
        img_url   = download_image("test")

        # go to part's "used with" page and fetch list of relevant products
        goto_used_with(sheet)
        used_with = fetch_used_with()

        # insert data to worksheet
        ws.cell(row = i, column = col["number"]).value    = number
        ws.cell(row = i, column = col["type"]).value      = part_type
        ws.cell(row = i, column = col["name"]).value      = name
        ws.cell(row = i, column = col["info"]).value      = info
        ws.cell(row = i, column = col["image url"]).value = img_url
        ws.cell(row = i, column = col["sheet"]).value     = sheet
        ws.cell(row = i, column = col["used with"]).value = used_with


    # save workbook
    wb.save(xlsx_file)

    # close browser
    print("Complete.\n")
    browser.close()
